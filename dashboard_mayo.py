"""
Dashboard de Cobranza Mayo 2025 — NAtura
Indicadores ejecutivos para reunion de cierre de mes
"""

import re
import random
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

random.seed(2025)
np.random.seed(2025)

# ─────────────────────────────────────────────
# CONFIGURACION
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Cobranza Mayo 2025 — NAtura",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─────────────────────────────────────────────
# PANTALLA DE INICIO
# ─────────────────────────────────────────────
if "archivos_listos" not in st.session_state:
    st.session_state.archivos_listos = False

DARK  = "#ffffff"
CARD  = "#f1f5f9"
BLUE  = "#2563eb"
GREEN = "#16a34a"
RED   = "#dc2626"
AMBER = "#d97706"
PURPLE= "#7c3aed"
SLATE = "#475569"
TEXT  = "#0f172a"
MUTED = "#64748b"

st.markdown(f"""
<style>
  .stApp {{ background-color: #ffffff; color: {TEXT}; }}
  [data-testid="stAppViewContainer"] {{ background-color: #ffffff; }}
  [data-testid="stHeader"] {{ background-color: #ffffff; }}
  [data-testid="stSidebar"] {{ background-color: #f8fafc; }}

  .stTabs [data-baseweb="tab-list"] {{
      background: #f1f5f9; border-radius: 10px; padding: 4px; gap: 4px;
  }}
  .stTabs [data-baseweb="tab"] {{
      color: {MUTED}; border-radius: 8px; font-weight: 600; font-size: 0.85rem;
  }}
  .stTabs [aria-selected="true"] {{
      background: {BLUE} !important; color: white !important;
  }}

  [data-testid="stDataFrame"] {{ border-radius: 10px; }}

  [data-testid="stMetric"] {{
      background: #f8fafc; border-radius: 10px; padding: 14px 18px;
      border-left: 3px solid {BLUE}; border: 1px solid #e2e8f0;
      border-left: 3px solid {BLUE};
  }}
  [data-testid="stMetricValue"] {{ color: {TEXT} !important; font-size: 1.8rem !important; }}
  [data-testid="stMetricLabel"] {{ color: {MUTED} !important; }}

  .sec {{ font-size: 1.2rem; font-weight: 700; color: {TEXT};
          border-bottom: 2px solid #e2e8f0; padding-bottom: 6px; margin: 8px 0 16px 0; }}

  .badge {{
      display:inline-block; padding: 2px 10px; border-radius: 20px;
      font-size: 0.78rem; font-weight: 700;
  }}
  .badge-red   {{ background:#fef2f2; color:{RED};   border:1px solid {RED}; }}
  .badge-green {{ background:#f0fdf4; color:{GREEN}; border:1px solid {GREEN}; }}
  .badge-amber {{ background:#fffbeb; color:{AMBER}; border:1px solid {AMBER}; }}

  [data-testid="stAlert"] {{ border-radius: 10px; }}
</style>
""", unsafe_allow_html=True)

PLOTLY_LAYOUT = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(color=TEXT, family="Inter, sans-serif"),
    margin=dict(l=10, r=10, t=30, b=10),
)

def apply_layout(fig, **kwargs):
    layout = {**PLOTLY_LAYOUT, **kwargs}
    for ax in ("xaxis", "yaxis", "xaxis2", "yaxis2"):
        base = dict(gridcolor="#e2e8f0", linecolor="#cbd5e1")
        if ax in kwargs:
            base.update(kwargs[ax])
            layout[ax] = base
    fig.update_layout(**layout)
    return fig

# ─────────────────────────────────────────────
# HELPERS DE CARGA
# ─────────────────────────────────────────────
def leer_archivo(f, cols_necesarias=None):
    if f is None:
        return None
    try:
        f.seek(0)
        if f.name.lower().endswith((".xlsx", ".xls")):
            # Leer solo encabezados primero
            df_head = pd.read_excel(f, nrows=0)
            f.seek(0)
            if cols_necesarias:
                cols_ok = [c for c in cols_necesarias if c in df_head.columns]
                if cols_ok:
                    return pd.read_excel(f, usecols=cols_ok, engine="openpyxl")
            return pd.read_excel(f, engine="openpyxl")
        else:
            df_head = pd.read_csv(f, nrows=0, encoding="utf-8-sig")
            f.seek(0)
            if cols_necesarias:
                cols_ok = [c for c in cols_necesarias if c in df_head.columns]
                if cols_ok:
                    return pd.read_csv(f, usecols=cols_ok, encoding="utf-8-sig", low_memory=False)
            return pd.read_csv(f, encoding="utf-8-sig", low_memory=False)
    except Exception:
        try:
            f.seek(0)
            return pd.read_csv(f, encoding="latin-1", low_memory=False)
        except Exception:
            return None

def col(df, *opciones):
    """Devuelve la primera columna que exista en df."""
    for o in opciones:
        if o in df.columns:
            return o
    return None

# ─────────────────────────────────────────────
# PLAN DE ACCIÓN COBRANZA — helpers de cálculo
# ─────────────────────────────────────────────
PAC_NAVY="#0D1B2A"; PAC_TEAL="#1B6CA8"; PAC_TEAL2="#2A9D8F"
PAC_AMBER="#E9C46A"; PAC_CORAL="#E76F51"; PAC_RED="#C1121F"

PAC_COLS_REQUERIDAS = [
    ("id","folio"),
    ("aging_de_morosidad",),
    ("dias_de_morosidad",),
    ("saldo_insoluto",),
    ("valor_original_deuda",),
    ("pago_actual",),
    ("status_rep",),
    ("edad_consultora",),
    ("segmentacion_rep",),
    ("division",),
    ("zona",),
    ("campana_numero",),
    ("score_riesgo",),
    ("numero_telefono_celular",),
    ("correo_electronico",),
]

PAC_UMBRALES = {"T1":30,"T2":60,"T3":90,"T4":120,"T5":150,"T6":180,"T7":None}

PAC_PAGO_LETRA = "CF"  # columna de Excel donde vive el pago (independiente del encabezado)
PAC_SALDO_LETRA = "I"  # columna de Excel donde vive el valor_saldo_deuda (independiente del encabezado)
PAC_ESTADO_GEO_LETRA = "AN"  # columna de Excel donde vive direccion_de_residencia_estado (independiente del encabezado)

def pac_col_por_letra(df, letra=PAC_PAGO_LETRA):
    """Devuelve el nombre de la columna del DataFrame en la posición de la letra de Excel dada."""
    idx = column_index_from_string(letra) - 1
    if 0 <= idx < len(df.columns):
        return df.columns[idx]
    return None

def pac_validar_columnas(df, pago_letra=PAC_PAGO_LETRA, saldo_letra=PAC_SALDO_LETRA):
    faltantes=[]
    encontradas={}
    cols_lower={c.lower().strip():c for c in df.columns}
    col_pago_letra = pac_col_por_letra(df, pago_letra)
    col_saldo_letra = pac_col_por_letra(df, saldo_letra)
    for opciones in PAC_COLS_REQUERIDAS:
        clave = opciones[0]
        if clave == "pago_actual" and col_pago_letra is not None:
            encontradas[clave] = col_pago_letra
            continue
        if clave == "saldo_insoluto" and col_saldo_letra is not None:
            encontradas[clave] = col_saldo_letra
            continue
        encontrada=None
        for o in opciones:
            if o in cols_lower:
                encontrada=cols_lower[o]; break
        if encontrada:
            encontradas[clave]=encontrada
        else:
            faltantes.append(" / ".join(opciones))
    return encontradas, faltantes

def pac_temporalidad(dias):
    if dias<=30: return "T1"
    elif dias<=60: return "T2"
    elif dias<=90: return "T3"
    elif dias<=120: return "T4"
    elif dias<=150: return "T5"
    elif dias<=180: return "T6"
    else: return "T7"

def pac_dias_para_migrar(dias, temporalidad):
    umbral=PAC_UMBRALES[temporalidad]
    if umbral is None:
        return 9999
    return max(umbral-dias,0)

def pac_riesgo(dias_para_migrar):
    if dias_para_migrar<=7: return "Crítico"
    elif dias_para_migrar<=15: return "Alto"
    elif dias_para_migrar<=30: return "Preventivo"
    else: return "Estable"

def pac_procesar(df_raw, encontradas):
    df=df_raw.copy()
    c=encontradas
    df[c["dias_de_morosidad"]]=pd.to_numeric(df[c["dias_de_morosidad"]],errors="coerce").fillna(0)
    df[c["saldo_insoluto"]]=pd.to_numeric(df[c["saldo_insoluto"]],errors="coerce").fillna(0)
    df[c["valor_original_deuda"]]=pd.to_numeric(df[c["valor_original_deuda"]],errors="coerce").fillna(0)
    df[c["pago_actual"]]=pd.to_numeric(df[c["pago_actual"]],errors="coerce").fillna(0)
    df[c["edad_consultora"]]=pd.to_numeric(df[c["edad_consultora"]],errors="coerce").fillna(0)
    df[c["score_riesgo"]]=pd.to_numeric(df[c["score_riesgo"]],errors="coerce").fillna(0)

    df["Temporalidad"]=df[c["dias_de_morosidad"]].apply(pac_temporalidad)
    df["DiasParaMigrar"]=df.apply(lambda r: pac_dias_para_migrar(r[c["dias_de_morosidad"]], r["Temporalidad"]), axis=1)
    df["RiesgoMigracion"]=df["DiasParaMigrar"].apply(pac_riesgo)

    bins=[0,30,45,60,200]
    labels=["18–30","31–45","46–60","61+"]
    df["RangoEdad"]=pd.cut(df[c["edad_consultora"]],bins=bins,labels=labels,right=True,include_lowest=True)
    df["RangoEdad"]=df["RangoEdad"].astype(str).replace("nan","Sin dato")

    score_bins=[-1,20,40,60,80,1000]
    score_labels=["Bajísimo","Bajo","Medio","Alto","Crítico"]
    df["ScoreCategoria"]=pd.cut(df[c["score_riesgo"]],bins=score_bins,labels=score_labels)

    return df

# ─────────────────────────────────────────────
# SEGUIMIENTO DIARIO DE RECUPERACIÓN — módulo independiente
# (no requiere subir la cartera completa del Plan de Acción Cobranza;
#  solo id/folio + pago (columna CF). Las dimensiones de desglose
#  -temporalidad, segmentación, estado, edad- son opcionales según lo
#  que traiga el archivo de cada día.)
# ─────────────────────────────────────────────
SEG_COLS_OPCIONALES = {
    "dias_de_morosidad": ("dias_de_morosidad","aging_de_morosidad"),
    "segmentacion_rep":  ("segmentacion_rep",),
    "status_rep":        ("status_rep",),
    "edad_consultora":   ("edad_consultora",),
    "saldo_insoluto":    ("saldo_insoluto",),
}

def seg_validar_columnas(df, pago_letra=PAC_PAGO_LETRA):
    cols_lower={c.lower().strip():c for c in df.columns}
    encontradas={}
    faltantes=[]
    id_col=None
    for o in ("id","folio"):
        if o in cols_lower:
            id_col=cols_lower[o]; break
    if id_col is None:
        faltantes.append("id / folio")
    else:
        encontradas["id"]=id_col

    col_pago = pac_col_por_letra(df, pago_letra)
    if col_pago is None:
        faltantes.append(f"pago (columna {pago_letra})")
    else:
        encontradas["pago_actual"]=col_pago

    for clave, opciones in SEG_COLS_OPCIONALES.items():
        for o in opciones:
            if o in cols_lower:
                encontradas[clave]=cols_lower[o]; break
    return encontradas, faltantes

def seg_procesar(df_raw, encontradas):
    df=df_raw.copy()
    c=encontradas
    df[c["pago_actual"]]=pd.to_numeric(df[c["pago_actual"]],errors="coerce").fillna(0)
    if "dias_de_morosidad" in c:
        dias=pd.to_numeric(df[c["dias_de_morosidad"]],errors="coerce").fillna(0)
        df["Temporalidad"]=dias.apply(pac_temporalidad)
        df["DiasParaMigrar"]=df.apply(lambda r: pac_dias_para_migrar(dias[r.name], r["Temporalidad"]),axis=1)
        df["RiesgoMigracion"]=df["DiasParaMigrar"].apply(pac_riesgo)
    if "edad_consultora" in c:
        edad=pd.to_numeric(df[c["edad_consultora"]],errors="coerce").fillna(0)
        df["RangoEdad"]=pd.cut(edad,bins=[0,30,45,60,200],labels=["18–30","31–45","46–60","61+"],right=True,include_lowest=True)
        df["RangoEdad"]=df["RangoEdad"].astype(str).replace("nan","Sin dato")
    if "saldo_insoluto" in c:
        df[c["saldo_insoluto"]]=pd.to_numeric(df[c["saldo_insoluto"]],errors="coerce").fillna(0)
    return df

def seg_recuperacion(snap_prev, snap_cur):
    """Compara dos snapshots livianos {fecha, df, cols} y calcula la recuperación
    generada entre la fecha anterior y la actual, cruzando por id."""
    df_p, c_p = snap_prev["df"], snap_prev["cols"]
    df_c, c_c = snap_cur["df"], snap_cur["cols"]

    cols_izq=[c_c["id"], c_c["pago_actual"]]
    rename_izq={c_c["id"]:"id", c_c["pago_actual"]:"pago_cur"}
    for clave in ["Temporalidad","RiesgoMigracion","RangoEdad"]:
        if clave in df_c.columns:
            cols_izq.append(clave); rename_izq[clave]=clave
    if "segmentacion_rep" in c_c:
        cols_izq.append(c_c["segmentacion_rep"]); rename_izq[c_c["segmentacion_rep"]]="Segmentacion"
    col_estado_geo = pac_col_por_letra(df_c, PAC_ESTADO_GEO_LETRA)
    if col_estado_geo is not None:
        cols_izq.append(col_estado_geo); rename_izq[col_estado_geo]="Estado"

    izq = df_c[cols_izq].copy().rename(columns=rename_izq)
    der = df_p[[c_p["id"], c_p["pago_actual"]]].copy().rename(columns={c_p["id"]:"id", c_p["pago_actual"]:"pago_prev"})

    merged = izq.merge(der, on="id", how="left")
    merged["pago_prev"] = merged["pago_prev"].fillna(0)
    merged["Recuperado"] = (merged["pago_cur"] - merged["pago_prev"]).clip(lower=0)
    return merged

def seg_resumen_por(merged, dim):
    if dim not in merged.columns:
        return None
    g = merged.groupby(dim).agg(
        Recuperado=("Recuperado","sum"),
        Consultoras=("id","count"),
        ConRecuperacion=("Recuperado", lambda s: (s>0).sum()),
    ).reset_index()
    g["% con recuperación"] = (g["ConRecuperacion"]/g["Consultoras"]*100).round(1)
    return g.sort_values("Recuperado", ascending=False)

def pac_export_html(df, encontradas):
    import json
    c=encontradas
    total_consultoras=len(df)
    saldo_total=float(df[c["saldo_insoluto"]].sum())
    deuda_total=float(df[c["valor_original_deuda"]].sum())
    pagos_total=float(df[c["pago_actual"]].sum())
    pct_recuperacion=(pagos_total/deuda_total*100) if deuda_total else 0.0
    criticas=df[df["RiesgoMigracion"]=="Crítico"]
    n_criticas=len(criticas)
    saldo_riesgo=float(criticas[c["saldo_insoluto"]].sum())

    orden_t=["T1","T2","T3","T4","T5","T6","T7"]
    saldo_por_t=df.groupby("Temporalidad")[c["saldo_insoluto"]].sum().reindex(orden_t).fillna(0)
    cuentas_por_t=df["Temporalidad"].value_counts().reindex(orden_t).fillna(0)
    orden_r=["Crítico","Alto","Preventivo","Estable"]
    riesgo_dist=df["RiesgoMigracion"].value_counts().reindex(orden_r).fillna(0)
    orden_edad=["18–30","31–45","46–60","61+"]
    edad_dist=df["RangoEdad"].value_counts().reindex(orden_edad).fillna(0)
    seg_saldo=df.groupby(c["segmentacion_rep"])[c["saldo_insoluto"]].sum().sort_values(ascending=False)

    fecha_str=pd.Timestamp.now().strftime("%d/%m/%Y %H:%M")

    html=f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<title>Plan de Acción Cobranza — NAtura</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1"></script>
<style>
  body{{font-family:'Segoe UI',system-ui,sans-serif;background:#f8fafc;color:{PAC_NAVY};margin:0;padding:24px;}}
  h1{{font-size:1.6rem;margin-bottom:0;}}
  .muted{{color:#64748b;font-size:0.85rem;}}
  .kpis{{display:flex;flex-wrap:wrap;gap:14px;margin:20px 0;}}
  .kpi{{background:#fff;border-radius:10px;box-shadow:0 1px 4px rgba(0,0,0,0.08);padding:16px 20px;flex:1;min-width:180px;border-left:4px solid {PAC_TEAL};}}
  .kpi b{{display:block;font-size:1.4rem;color:{PAC_NAVY};}}
  .charts{{display:flex;flex-wrap:wrap;gap:20px;margin-top:20px;}}
  .chart-box{{background:#fff;border-radius:10px;box-shadow:0 1px 4px rgba(0,0,0,0.08);padding:16px;flex:1;min-width:320px;}}
  @media(max-width:700px){{.kpis,.charts{{flex-direction:column;}}}}
</style></head>
<body>
  <h1>📋 Plan de Acción Cobranza — NAtura</h1>
  <div class="muted">Reporte generado el {fecha_str}</div>
  <div class="kpis">
    <div class="kpi">Consultoras en cartera<b>{total_consultoras:,}</b></div>
    <div class="kpi">Deuda asignada<b>${saldo_total:,.0f}</b></div>
    <div class="kpi">Deuda original total<b>${deuda_total:,.0f}</b></div>
    <div class="kpi">Pagos registrados<b>${pagos_total:,.0f}</b></div>
    <div class="kpi">% Recuperación<b>{pct_recuperacion:.1f}%</b></div>
    <div class="kpi">Migran en ≤7 días<b>{n_criticas:,} (${saldo_riesgo:,.0f})</b></div>
  </div>
  <div class="charts">
    <div class="chart-box"><canvas id="cSaldoT"></canvas></div>
    <div class="chart-box"><canvas id="cRiesgo"></canvas></div>
    <div class="chart-box"><canvas id="cCuentasT"></canvas></div>
    <div class="chart-box"><canvas id="cEdad"></canvas></div>
    <div class="chart-box"><canvas id="cSeg"></canvas></div>
  </div>
<script>
const COLORS={{navy:"{PAC_NAVY}",teal:"{PAC_TEAL}",teal2:"{PAC_TEAL2}",amber:"{PAC_AMBER}",coral:"{PAC_CORAL}",red:"{PAC_RED}"}};
new Chart(document.getElementById('cSaldoT'),{{type:'bar',data:{{labels:{json.dumps(orden_t)},
  datasets:[{{label:'Saldo insoluto por temporalidad',data:{json.dumps([round(v,2) for v in saldo_por_t.tolist()])},backgroundColor:COLORS.teal}}]}},
  options:{{plugins:{{title:{{display:true,text:'Saldo por Temporalidad'}}}}}}}});
new Chart(document.getElementById('cRiesgo'),{{type:'doughnut',data:{{labels:{json.dumps(orden_r)},
  datasets:[{{data:{json.dumps([int(v) for v in riesgo_dist.tolist()])},backgroundColor:[COLORS.red,COLORS.coral,COLORS.amber,COLORS.teal2]}}]}},
  options:{{plugins:{{title:{{display:true,text:'Distribución de Riesgo de Migración'}}}}}}}});
new Chart(document.getElementById('cCuentasT'),{{type:'bar',data:{{labels:{json.dumps(orden_t)},
  datasets:[{{label:'Consultoras por temporalidad',data:{json.dumps([int(v) for v in cuentas_por_t.tolist()])},backgroundColor:COLORS.navy}}]}},
  options:{{plugins:{{title:{{display:true,text:'Consultoras por Temporalidad'}}}}}}}});
new Chart(document.getElementById('cEdad'),{{type:'doughnut',data:{{labels:{json.dumps(orden_edad)},
  datasets:[{{data:{json.dumps([int(v) for v in edad_dist.tolist()])},backgroundColor:[COLORS.teal,COLORS.teal2,COLORS.amber,COLORS.coral]}}]}},
  options:{{plugins:{{title:{{display:true,text:'Distribución por Edad'}}}}}}}});
new Chart(document.getElementById('cSeg'),{{type:'bar',data:{{labels:{json.dumps(seg_saldo.index.tolist())},
  datasets:[{{label:'Saldo por segmentación',data:{json.dumps([round(v,2) for v in seg_saldo.tolist()])},backgroundColor:COLORS.coral}}]}},
  options:{{indexAxis:'y',plugins:{{title:{{display:true,text:'Saldo por Segmentación'}}}}}}}});
</script>
</body></html>"""
    return html

# ─────────────────────────────────────────────
# DATOS SINTETICOS MAYO 2025 (fallback)
# ─────────────────────────────────────────────
ESTADOS   = ["CDMX","Edo. de México","Jalisco","Nuevo León","Puebla",
             "Guanajuato","Veracruz","Michoacán","Chihuahua","Tamaulipas"]
SECTORES  = ["Belleza","Moda","Hogar","Nutrición","Bienestar","Joyería"]
SEGMENTOS = ["Diamante","Oro","Plata","Bronce","Nuevo Ingreso"]
GVS       = [f"GV-{str(i).zfill(3)}" for i in range(1, 16)]
ASESORES  = [f"Asesor {chr(65+i)}" for i in range(10)]
HORAS     = list(range(8, 21))

# ── Cierre de mes ────────────────────────────
META_TOTAL       = 6_000_000
RECUPERADO_TOTAL = 4_954_931.88
PROMESAS_GEN     = 1_240
PROMESAS_CUMP    = 748
PROMESAS_CAIDAS  = PROMESAS_GEN - PROMESAS_CUMP

dias = pd.date_range("2025-05-01", "2025-05-31", freq="B")
meta_d = META_TOTAL / len(dias)
rec_d = np.clip(np.random.normal(meta_d * 0.92, meta_d * 0.15, len(dias)), meta_d * 0.4, meta_d * 1.3)
df_cierre = pd.DataFrame({
    "Fecha": dias,
    "Meta": meta_d,
    "Recuperado": rec_d,
})
df_cierre["Meta Acum"]       = df_cierre["Meta"].cumsum()
df_cierre["Recuperado Acum"] = df_cierre["Recuperado"].cumsum()

df_motivos = pd.DataFrame({
    "Motivo": ["Sin liquidez","No contestó","Negó deuda","Promesa vencida","Número inválido","Otro"],
    "Casos":  [198, 167, 89, 134, 72, 61],
})

# ── Contactabilidad ──────────────────────────
TOTAL_LLAMADAS = 38_420   # total filas en gestiones (se sobreescribe con archivo real)
TOTAL_INT  = 38_420
TITULAR    = 14_820
BUZON      = 9_105
NO_CONT    = 8_934
NUM_INV    = 3_210
COLGADO    = 2_351
CR         = TITULAR / TOTAL_INT

df_horario = pd.DataFrame({
    "Hora": HORAS,
    "Contactos": [120,310,520,680,740,810,860,780,690,620,480,310,210],
    "Promesas":  [ 12, 35, 62, 88, 95,104,112,100, 89, 78, 61, 39, 25],
})
df_horario["Conv %"] = (df_horario["Promesas"] / df_horario["Contactos"] * 100).round(1)

df_canal = pd.DataFrame({
    "Canal":      ["SMS","WhatsApp","Email","Llamada","IVR"],
    "Enviados":   [12400, 8300, 9800, 38420, 4200],
    "Respuestas": [1240,  2076,  588, 14820,  840],
    "Promesas":   [ 186,   622,   74,   748,  168],
})
df_canal["Tasa resp %"] = (df_canal["Respuestas"] / df_canal["Enviados"] * 100).round(1)
df_canal["Conv prom %"] = (df_canal["Promesas"]   / df_canal["Respuestas"] * 100).round(1)

# ── Indicadores ──────────────────────────────
df_sector = pd.DataFrame({
    "Sector": SECTORES,
    "Meta":       [980000, 870000, 650000, 720000, 540000, 490000],
    "Recuperado": [784000, 652500, 487500, 504000, 378000, 318000],
})
df_sector["Cumplimiento %"] = (df_sector["Recuperado"] / df_sector["Meta"] * 100).round(1)

np.random.seed(42)
df_gv = pd.DataFrame({
    "GV":         GVS,
    "Meta":       np.random.randint(200000, 500000, len(GVS)),
    "Recuperado": np.random.randint(140000, 460000, len(GVS)),
})
df_gv["Cumplimiento %"] = (df_gv["Recuperado"] / df_gv["Meta"] * 100).round(1)
df_gv = df_gv.sort_values("Cumplimiento %", ascending=False).reset_index(drop=True)

df_edad = pd.DataFrame({
    "Rango":     ["18-25","26-35","36-45","46-55","56-65","66+"],
    "Cuentas":   [1240, 3870, 4210, 3650, 2180, 980],
    "Recuperado":[74400,271400,336800,255500,130800,49000],
    "Promesas %":[38,62,71,65,52,41],
})
df_edad["Ticket Prom"] = (df_edad["Recuperado"] / df_edad["Cuentas"]).round(0)

df_estado = pd.DataFrame({
    "Estado":     ESTADOS,
    "Cuentas":    [4820,3610,2840,1980,1650,1420,1280,1190,980,870],
    "Recuperado": [1120000,892000,620000,485000,378000,312000,280000,245000,198000,168000],
})
df_estado["Ticket Prom"] = (df_estado["Recuperado"] / df_estado["Cuentas"]).round(0)

df_segmento = pd.DataFrame({
    "Segmento":      SEGMENTOS,
    "Cuentas":       [890,2340,4210,5680,2320],
    "Recuperado":    [534000,936000,1263000,1022400,139200],
    "Cumplimiento %":[87,82,74,68,45],
})

# ── Operación ────────────────────────────────
np.random.seed(99)
df_asesores = pd.DataFrame({
    "Asesor":            ASESORES,
    "Llamadas":          np.random.randint(280, 520, 10),
    "TMO (min)":         np.random.uniform(3.2, 6.8, 10).round(1),
    "Contactos Titular": np.random.randint(80, 180, 10),
    "Promesas":          np.random.randint(20, 65, 10),
    "Colgadas <30s":     np.random.randint(15, 60, 10),
    "Monto Rec":         np.random.randint(180000, 520000, 10),
})
df_asesores["% Contacto"] = (df_asesores["Contactos Titular"] / df_asesores["Llamadas"] * 100).round(1)
df_asesores["% Abandono"] = (df_asesores["Colgadas <30s"]     / df_asesores["Llamadas"] * 100).round(1)
df_asesores["Conv %"]     = (df_asesores["Promesas"]           / df_asesores["Contactos Titular"] * 100).round(1)
df_asesores = df_asesores.sort_values("Monto Rec", ascending=False).reset_index(drop=True)

df_objeciones = pd.DataFrame({
    "Objeción":           ["No tengo dinero","No reconozco la deuda","Ya pagué",
                           "Espera a quincena","Mándame estado de cuenta","No soy yo"],
    "Frecuencia":         [2840, 1920, 1480, 2210, 890, 640],
    "Resolución exitosa %":[38,  51,   72,   61,   44,  58],
})

# ── Plan de trabajo ──────────────────────────
META_JUN      = 5_200_000
PROY_JUN      = 4_420_000
PIPELINE_PROM = 1_840_000

dias_jun = pd.date_range("2025-06-01","2025-06-30", freq="B")
n = len(dias_jun)
df_proy = pd.DataFrame({
    "Fecha":      dias_jun,
    "Meta":       np.linspace(0, META_JUN,        n),
    "Base":       np.linspace(0, PROY_JUN,        n),
    "Optimista":  np.linspace(0, META_JUN * 1.05, n),
    "Pesimista":  np.linspace(0, PROY_JUN * 0.88, n),
})

df_palancas = pd.DataFrame({
    "Palanca":             ["Recuperar promesas caídas Mayo","Incrementar contact rate",
                            "Mejorar conv. WhatsApp","Script objeciones","GVs rezagados"],
    "Impacto estimado $":  [420000, 280000, 180000, 150000, 220000],
})

df_acciones = pd.DataFrame([
    ("Inmediato","Operaciones","Rellamada a 2,891 promesas caídas"),
    ("Inmediato","Analítica","Reporte GV diario automatizado"),
    ("Lun 9 Jun","Canales Digitales","Lanzar A/B test SMS por segmento"),
    ("Mar 10 Jun","Supervisión","Coaching script objeciones 'sin dinero'"),
    ("Jue 12 Jun","Gerencia","Validar incentivo para GVs rezagados"),
    ("Vie 13 Jun","Operaciones","Ajuste de marcaciones franja 12-15h +25%"),
], columns=["Plazo","Área","Acción"])

# ─────────────────────────────────────────────
# PANTALLA DE INICIO — CARGA DE ARCHIVOS
# ─────────────────────────────────────────────
if not st.session_state.archivos_listos:
    st.markdown("""
    <div style="text-align:center; padding: 40px 0 20px 0;">
        <div style="font-size:4rem;">📊</div>
        <h1 style="font-size:2.2rem; font-weight:800; color:#0f172a; margin:10px 0 4px 0;">
            Dashboard de Cobranza
        </h1>
        <h2 style="font-size:1.4rem; font-weight:400; color:#64748b; margin:0 0 8px 0;">
            Mayo 2025 — NAtura
        </h2>
        <p style="color:#94a3b8; font-size:0.95rem;">
            Sube tus archivos para generar los indicadores con datos reales.<br>
            También puedes continuar con datos de demostración.
        </p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("---")

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("#### 📋 Cartera / Remesa")
        f_cartera = st.file_uploader("Archivo de cartera asignada", type=["csv","xlsx","xls"], key="cartera", label_visibility="collapsed")
        st.caption("Columnas: `valor_saldo_deuda`, `zona`, `estado`, `segmento_nombre`, `edad_consultora`")

        st.markdown("#### 💰 Pagos / Recuperación")
        f_pagos = st.file_uploader("Archivo de pagos", type=["csv","xlsx","xls"], key="pagos", label_visibility="collapsed")
        st.caption("Columnas: `monto_pagado`, `fecha_pago`, `asesor`")

    with col2:
        st.markdown("#### 📞 Gestión de llamadas")
        f_gestion = st.file_uploader("Archivo de gestión", type=["csv","xlsx","xls"], key="gestion", label_visibility="collapsed")
        st.caption("Columnas: `resultado`, `hora_llamada`, `canal`, `asesor`, `duracion_seg`")

        st.markdown("#### 🤝 Promesas de pago")
        f_promesas = st.file_uploader("Archivo de promesas", type=["csv","xlsx","xls"], key="promesas", label_visibility="collapsed")
        st.caption("Columnas: `monto_promesa`, `fecha_promesa`, `cumplida`")

        st.markdown("#### 📑 Comparativo Cobranza")
        f_comparativo = st.file_uploader("Archivo Comparativo (Pagos + Comparativo)", type=["csv","xlsx","xls"], key="comparativo", label_visibility="collapsed")
        st.caption("Hojas: `TABLA DE PAGOS` (filas T1..Tn + TOTAL), `COMPARATIVO` (Cartera, Objetivo, Cobranza, Tramo)")

    st.markdown("---")
    col_btn1, col_btn2, col_btn3 = st.columns([2, 1, 2])
    with col_btn2:
        archivos_subidos = any([f_cartera, f_pagos, f_gestion, f_promesas, f_comparativo])
        btn_label = "▶ Ver Dashboard con mis datos" if archivos_subidos else "▶ Ver Dashboard de demostración"
        if st.button(btn_label, type="primary", use_container_width=True):
            st.session_state.archivos_listos = True
            st.session_state.f_cartera  = f_cartera
            st.session_state.f_pagos    = f_pagos
            st.session_state.f_gestion  = f_gestion
            st.session_state.f_promesas = f_promesas
            st.session_state.f_comparativo = f_comparativo
            st.rerun()
    st.stop()

# ─────────────────────────────────────────────
# RECUPERAR ARCHIVOS DE SESSION STATE
# ─────────────────────────────────────────────
f_cartera  = st.session_state.get("f_cartera")
f_pagos    = st.session_state.get("f_pagos")
f_gestion  = st.session_state.get("f_gestion")
f_promesas = st.session_state.get("f_promesas")
f_comparativo = st.session_state.get("f_comparativo")

with st.sidebar:
    MODO_REAL = any([f_cartera, f_pagos, f_gestion, f_promesas, f_comparativo])
    if MODO_REAL:
        st.success("✅ Datos reales activos")
        for nombre, f in [("Cartera",f_cartera),("Pagos",f_pagos),
                           ("Gestión",f_gestion),("Promesas",f_promesas),
                           ("Comparativo",f_comparativo)]:
            if f:
                df_tmp = leer_archivo(f)
                if df_tmp is not None:
                    st.markdown(f"**{nombre}** — {len(df_tmp):,} filas")
    else:
        st.warning("⚠️ Modo demostración")
    st.markdown("---")
    if st.button("← Cambiar archivos"):
        st.session_state.archivos_listos = False
        st.rerun()

# ─────────────────────────────────────────────
# CARGAR DATAFRAMES REALES SI EXISTEN
# ─────────────────────────────────────────────
COLS_CARTERA  = ["codigo de cliente","aging_de_morosidad","valor_saldo_deuda",
                  "zona","estado","segmento_nombre","edad_consultora",
                  "division","red","direccion_de_residencia_estado","numero_clave",
                  "L","Saldo","BU (Division)","BF","Division"]
COLS_PAGOS    = ["BU (Pago total)","codigo de cliente","monto_pagado","valor_pago",
                  "importe","monto","pago","fecha_pago","fecha","asesor","ejecutivo",
                  "zona","direccion_de_residencia_estado","estado","edad_consultora",
                  "segmento_nombre","aging_de_morosidad"]
COLS_GESTION  = ["codigo de cliente","medicion","K","hora_llamada","hora",
                  "canal","medio","asesor","ejecutivo","duracion_seg","duracion","tmo"]
COLS_PROMESAS = ["monto_promesa","promesa_monto","monto","importe","codigo de cliente",
                  "fecha_promesa","fecha","cumplida","estatus","status","resultado"]

@st.cache_data(show_spinner="Cargando archivos, espera...")
def cargar_desde_bytes(bytes_data, nombre, cols):
    import io
    f = io.BytesIO(bytes_data)
    try:
        if nombre.lower().endswith((".xlsx", ".xls")):
            df_head = pd.read_excel(f, nrows=0, engine="openpyxl")
            f.seek(0)
            cols_ok = [c for c in cols if c in df_head.columns]
            return pd.read_excel(f, usecols=cols_ok if cols_ok else None, engine="openpyxl")
        else:
            df_head = pd.read_csv(f, nrows=0, encoding="utf-8-sig")
            f.seek(0)
            cols_ok = [c for c in cols if c in df_head.columns]
            return pd.read_csv(f, usecols=cols_ok if cols_ok else None, encoding="utf-8-sig", low_memory=False)
    except Exception:
        try:
            f.seek(0)
            return pd.read_csv(f, encoding="latin-1", low_memory=False)
        except Exception:
            return None

def cargar_si_existe(f, cols):
    if f is None:
        return None
    return cargar_desde_bytes(f.read(), f.name, cols)

df_cart_real  = cargar_si_existe(f_cartera,  COLS_CARTERA)
df_pago_real  = cargar_si_existe(f_pagos,    COLS_PAGOS)
df_gest_real  = cargar_si_existe(f_gestion,  COLS_GESTION)
df_prom_real  = cargar_si_existe(f_promesas, COLS_PROMESAS)

# ── COMPARATIVO COBRANZA POR TRAMO (sincronización Pagos → Comparativo) ──
df_pagos_diario=None
df_comparativo=None
RESUMEN_COMP=None
DEBUG_COMP=None
if f_comparativo is not None:
    DEBUG_COMP={}
    try:
        wb=openpyxl.load_workbook(f_comparativo,data_only=True)
        DEBUG_COMP["hojas"]=wb.sheetnames
        sh_pagos=next((s for s in wb.sheetnames if "PAGO" in s.upper()),None)
        sh_comp =next((s for s in wb.sheetnames if "COMPARATIV" in s.upper()),None)
        DEBUG_COMP["hoja_pagos"]=sh_pagos
        DEBUG_COMP["hoja_comparativo"]=sh_comp
        if sh_pagos and sh_comp:
            ws_p=wb[sh_pagos]; ws_c=wb[sh_comp]

            # localizar fila de encabezados (T1..Tn) en hoja de pagos
            header_row=None
            for r in range(1,min(ws_p.max_row,10)+1):
                row_vals=[ws_p.cell(row=r,column=c).value for c in range(1,ws_p.max_column+1)]
                if any(isinstance(v,str) and v.strip().upper()=="T1" for v in row_vals):
                    header_row=r; break
            tramo_cols={}
            if header_row:
                for c in range(1,ws_p.max_column+1):
                    v=ws_p.cell(row=header_row,column=c).value
                    if isinstance(v,str):
                        vu=v.strip().upper()
                        if len(vu)>=2 and vu[0]=="T" and vu[1:].isdigit():
                            tramo_cols[vu]=c
            DEBUG_COMP["header_row_pagos"]=header_row
            DEBUG_COMP["tramo_cols"]=tramo_cols

            # localizar fila TOTAL
            total_row=None
            if header_row:
                for r in range(header_row+1,ws_p.max_row+1):
                    v1=ws_p.cell(row=r,column=1).value
                    v2=ws_p.cell(row=r,column=2).value
                    if (isinstance(v1,str) and "TOTAL" in v1.upper()) or (isinstance(v2,str) and "TOTAL" in v2.upper()):
                        total_row=r; break

            DEBUG_COMP["total_row_pagos"]=total_row

            # totales por tramo (fila TOTAL)
            totales_tramo={t:float(ws_p.cell(row=total_row,column=c).value or 0) for t,c in tramo_cols.items()} if total_row else {}
            DEBUG_COMP["totales_tramo"]=totales_tramo

            # pagos diarios (columna B = fecha, suma de tramos por día)
            registros=[]
            if header_row and total_row:
                for r in range(header_row+1,total_row):
                    fecha=ws_p.cell(row=r,column=2).value
                    if fecha is None: continue
                    total_dia=sum(float(ws_p.cell(row=r,column=c).value or 0) for c in tramo_cols.values())
                    registros.append({"Fecha":fecha,"Total Día":total_dia})
            if registros:
                df_pagos_diario=pd.DataFrame(registros)
                df_pagos_diario["Fecha"]=pd.to_datetime(df_pagos_diario["Fecha"],errors="coerce")
                df_pagos_diario=df_pagos_diario.dropna(subset=["Fecha"]).sort_values("Fecha").reset_index(drop=True)
                df_pagos_diario["Acumulado"]=df_pagos_diario["Total Día"].cumsum()

            # localizar encabezados en hoja Comparativo
            header_row_c=None
            for r in range(1,min(ws_c.max_row,10)+1):
                row_vals=[ws_c.cell(row=r,column=c).value for c in range(1,ws_c.max_column+1)]
                if any(isinstance(v,str) and "CARTERA" in v.upper() for v in row_vals):
                    header_row_c=r; break

            DEBUG_COMP["header_row_comparativo"]=header_row_c

            cols_c={}
            if header_row_c:
                for c in range(1,ws_c.max_column+1):
                    v=ws_c.cell(row=header_row_c,column=c).value
                    if not isinstance(v,str): continue
                    vu=v.strip().upper()
                    if "CARTERA" in vu: cols_c["Cartera"]=c
                    elif "OBJETIVO" in vu and "%" in vu: cols_c["Objetivo %"]=c
                    elif "OBJETIVO" in vu: cols_c["Objetivo $"]=c
                    elif "CUENTA" in vu: cols_c["Cuentas"]=c
                    elif "COBRANZA" in vu: cols_c["Cobranza"]=c
                    elif "TRAMO" in vu or "TEMPORALIDAD" in vu: cols_c["Tramo"]=c
            DEBUG_COMP["cols_comparativo"]=cols_c

            # valores de la columna Tramo detectados (para comparar con totales_tramo)
            tramo_vals_c=[]
            if header_row_c:
                col_tramo_dbg=cols_c.get("Tramo",1)
                for r in range(header_row_c+1,ws_c.max_row+1):
                    v=ws_c.cell(row=r,column=col_tramo_dbg).value
                    if v is not None:
                        tramo_vals_c.append(v)
            DEBUG_COMP["tramo_vals_comparativo"]=tramo_vals_c

            # sincronizar Cobranza $ por tramo (escribe el valor calculado en la hoja en memoria)
            comp_rows=[]
            if header_row_c and "Cobranza" in cols_c:
                col_tramo=cols_c.get("Tramo",1)
                for r in range(header_row_c+1,ws_c.max_row+1):
                    tramo_val=ws_c.cell(row=r,column=col_tramo).value
                    if not isinstance(tramo_val,str): continue
                    m=re.search(r"\d+",tramo_val)
                    tramo_key=f"T{m.group(0)}" if m else tramo_val.strip().upper()
                    if tramo_key not in totales_tramo: continue
                    cobranza_val=totales_tramo[tramo_key]
                    ws_c.cell(row=r,column=cols_c["Cobranza"]).value=cobranza_val
                    comp_rows.append({
                        "Tramo":tramo_val.strip(),
                        "Cuentas":ws_c.cell(row=r,column=cols_c["Cuentas"]).value if "Cuentas" in cols_c else None,
                        "Cartera $":float(ws_c.cell(row=r,column=cols_c["Cartera"]).value or 0) if "Cartera" in cols_c else 0.0,
                        "Objetivo $":float(ws_c.cell(row=r,column=cols_c["Objetivo $"]).value or 0) if "Objetivo $" in cols_c else 0.0,
                        "Objetivo %":ws_c.cell(row=r,column=cols_c["Objetivo %"]).value if "Objetivo %" in cols_c else None,
                        "Cobranza $":cobranza_val,
                    })

            if comp_rows:
                df_comparativo=pd.DataFrame(comp_rows)
                df_comparativo["Cumplimiento %"]=(df_comparativo["Cobranza $"]/df_comparativo["Objetivo $"].replace(0,np.nan)*100).round(1)

                total_cartera =df_comparativo["Cartera $"].sum()
                total_objetivo=df_comparativo["Objetivo $"].sum()
                total_cobranza=df_comparativo["Cobranza $"].sum()

                if df_pagos_diario is not None and len(df_pagos_diario)>0:
                    ult_fecha=df_pagos_diario["Fecha"].max()
                    dias_transcurridos=df_pagos_diario["Fecha"].nunique()
                    dias_totales_mes=pd.Period(ult_fecha,freq="M").days_in_month
                    dias_pendientes=max(dias_totales_mes-ult_fecha.day,0)
                    promedio_diario=total_cobranza/max(dias_transcurridos,1)
                    proyeccion_total=total_cobranza+promedio_diario*dias_pendientes
                else:
                    ult_fecha=None; dias_transcurridos=0; dias_totales_mes=0
                    dias_pendientes=0; promedio_diario=0.0; proyeccion_total=total_cobranza

                RESUMEN_COMP=dict(
                    total_cartera=total_cartera,total_objetivo=total_objetivo,total_cobranza=total_cobranza,
                    cumplimiento=(total_cobranza/total_objetivo*100) if total_objetivo else 0.0,
                    dias_transcurridos=dias_transcurridos,dias_pendientes=dias_pendientes,
                    dias_totales_mes=dias_totales_mes,promedio_diario=promedio_diario,
                    proyeccion_total=proyeccion_total,ultima_fecha=ult_fecha,
                )
    except Exception as e:
        st.sidebar.error(f"Error procesando archivo Comparativo: {e}")
        DEBUG_COMP["error"]=str(e)

# ── Calcular División real si hay datos ──────
if df_cart_real is not None:
    st.sidebar.write("**Cols Remesa:**", list(df_cart_real.columns))
if df_pago_real is not None:
    st.sidebar.write("**Cols Pagos:**", list(df_pago_real.columns))
if df_cart_real is not None and df_pago_real is not None:
    c_div_rem  = col(df_cart_real, "BU (Division)","Division","division","BF")
    c_cod_rem  = col(df_cart_real, "codigo de cliente","codigo_de_cliente")
    c_meta_rem = col(df_cart_real, "L","Saldo","valor_saldo_deuda")
    c_cod_pag  = col(df_pago_real, "codigo de cliente","codigo_de_cliente")
    c_pago_bu  = col(df_pago_real, "BU (Pago total)","monto_pagado","valor_pago")

    if c_div_rem and c_cod_rem and c_meta_rem and c_cod_pag and c_pago_bu:
        df_rem_unico = df_cart_real.drop_duplicates(subset=[c_cod_rem]).copy()
        df_rem_unico[c_meta_rem] = pd.to_numeric(df_rem_unico[c_meta_rem], errors="coerce").fillna(0)
        asig_div = df_rem_unico.groupby(c_div_rem)[c_meta_rem].sum().reset_index()
        asig_div.columns = ["Division","Meta"]

        df_pago_real[c_pago_bu] = pd.to_numeric(df_pago_real[c_pago_bu], errors="coerce").fillna(0)
        df_cruce = df_pago_real[[c_cod_pag, c_pago_bu]].merge(
            df_rem_unico[[c_cod_rem, c_div_rem]],
            left_on=c_cod_pag, right_on=c_cod_rem, how="left"
        )
        pago_div = df_cruce.groupby(c_div_rem)[c_pago_bu].sum().reset_index()
        pago_div.columns = ["Division","Recuperado"]

        df_sector = asig_div.merge(pago_div, on="Division", how="left")
        df_sector["Recuperado"] = df_sector["Recuperado"].fillna(0)
        df_sector["Cumplimiento %"] = (df_sector["Recuperado"] / df_sector["Meta"].clip(1) * 100).round(1)
        df_sector = df_sector.rename(columns={"Division":"Sector"})

# ── Calcular métricas reales si hay datos ────
if df_pago_real is not None:
    c_monto = col(df_pago_real, "BU (Pago total)","monto_pagado","valor_pago","importe","monto","pago")
    c_fecha = col(df_pago_real, "fecha_pago","fecha","date","fecha_operacion")
    c_asesor= col(df_pago_real, "asesor","ejecutivo","agente","nombre_asesor")
    if c_monto:
        df_pago_real[c_monto] = pd.to_numeric(df_pago_real[c_monto], errors="coerce").fillna(0)
    RECUPERADO_TOTAL = df_pago_real[c_monto].sum() if c_monto else RECUPERADO_TOTAL
    if c_fecha and c_monto:
        df_pago_real[c_fecha] = pd.to_datetime(df_pago_real[c_fecha], errors="coerce")
        df_cierre = (df_pago_real.dropna(subset=[c_fecha])
                     .groupby(c_fecha)[c_monto].sum()
                     .reset_index()
                     .rename(columns={c_fecha:"Fecha", c_monto:"Recuperado"}))
        df_cierre = df_cierre[df_cierre["Fecha"].dt.month == 5].sort_values("Fecha")
        df_cierre["Meta"] = META_TOTAL / max(len(df_cierre), 1)
        df_cierre["Meta Acum"]       = df_cierre["Meta"].cumsum()
        df_cierre["Recuperado Acum"] = df_cierre["Recuperado"].cumsum()

if df_cart_real is not None:
    c_meta   = col(df_cart_real, "L","Saldo","valor_saldo_deuda")
    c_codigo = col(df_cart_real, "codigo de cliente","codigo_de_cliente","numero_clave")
    c_aging  = col(df_cart_real, "aging_de_morosidad","aging","temporalidad")
    if c_meta and c_codigo:
        df_cart_real[c_meta] = pd.to_numeric(df_cart_real[c_meta], errors="coerce").fillna(0)
        df_meta_unico = df_cart_real.drop_duplicates(subset=[c_codigo])
        META_TOTAL = df_meta_unico[c_meta].sum()

if df_prom_real is not None:
    c_mp  = col(df_prom_real, "monto_promesa","promesa_monto","monto","importe")
    c_cum = col(df_prom_real, "cumplida","estatus","status","resultado")
    PROMESAS_GEN   = len(df_prom_real)
    if c_cum:
        cumplidas = df_prom_real[c_cum].astype(str).str.lower().isin(["1","si","sí","cumplida","pagada","true"])
        PROMESAS_CUMP  = int(cumplidas.sum())
        PROMESAS_CAIDAS= PROMESAS_GEN - PROMESAS_CUMP
    if c_mp:
        PIPELINE_PROM = df_prom_real[c_mp].sum()

if df_gest_real is not None:
    c_res  = col(df_gest_real, "medicion","K","resultado","resultado_gestion","disposicion")
    c_cod  = col(df_gest_real, "codigo de cliente","codigo_de_cliente","numero_clave")
    c_hora = col(df_gest_real, "hora_llamada","hora","hour")
    c_dur  = col(df_gest_real, "duracion_seg","duracion","duration","tmo")
    c_canal= col(df_gest_real, "canal","medio","channel","tipo_contacto")
    c_as   = col(df_gest_real, "asesor","ejecutivo","agente")

    CONTACTADOS = ["contacto directo","contacto indirecto"]
    NO_CONTACTADO = ["no contactado"]

    TOTAL_LLAMADAS = len(df_gest_real)
    if c_res and c_cod:
        res_lower = df_gest_real[c_res].astype(str).str.strip().str.lower()
        # Contact Rate por código único
        df_gest_real["_contactado"] = res_lower.isin([c.lower() for c in CONTACTADOS])
        df_cod_unico = df_gest_real.groupby(c_cod)["_contactado"].max().reset_index()
        TITULAR   = int(df_cod_unico["_contactado"].sum())
        TOTAL_INT = len(df_cod_unico)
        NO_CONT   = TOTAL_INT - TITULAR
        BUZON     = 0
        NUM_INV   = 0
        COLGADO   = 0
        CR        = TITULAR / max(TOTAL_INT, 1)
    elif c_res:
        res_lower = df_gest_real[c_res].astype(str).str.strip().str.lower()
        TITULAR   = int(res_lower.isin([c.lower() for c in CONTACTADOS]).sum())
        TOTAL_INT = len(df_gest_real)
        NO_CONT   = TOTAL_INT - TITULAR
        CR        = TITULAR / max(TOTAL_INT, 1)
    if c_hora:
        try:
            df_gest_real["_hora"] = pd.to_datetime(df_gest_real[c_hora], errors="coerce").dt.hour
            if c_res:
                grp = df_gest_real.groupby("_hora")
                df_horario = pd.DataFrame({
                    "Hora":      grp.size().index.tolist(),
                    "Contactos": grp.size().values.tolist(),
                    "Promesas":  grp.apply(lambda x: x[c_res].astype(str).str.lower()
                                           .isin(["promesa","pdc","promesa de pago"]).sum()).values.tolist(),
                })
                df_horario["Conv %"] = (df_horario["Promesas"] / df_horario["Contactos"].clip(1) * 100).round(1)
        except Exception:
            pass
    if c_canal and c_res:
        def canal_stats(grp):
            total = len(grp)
            resp  = grp[c_res].notna().sum()
            prom  = grp[c_res].astype(str).str.lower().isin(["promesa","pdc","promesa de pago"]).sum()
            return pd.Series({"Enviados": total, "Respuestas": resp, "Promesas": prom})
        df_canal = df_gest_real.groupby(c_canal).apply(canal_stats).reset_index()
        df_canal.columns = ["Canal","Enviados","Respuestas","Promesas"]
        df_canal["Tasa resp %"] = (df_canal["Respuestas"] / df_canal["Enviados"].clip(1) * 100).round(1)
        df_canal["Conv prom %"] = (df_canal["Promesas"]   / df_canal["Respuestas"].clip(1) * 100).round(1)
    if c_as and c_res:
        def asesor_stats(grp):
            llamadas = len(grp)
            titular  = grp[c_res].astype(str).str.lower().isin(["contacto titular","titular","contactado","promesa","pdc"]).sum()
            promesas = grp[c_res].astype(str).str.lower().isin(["promesa","pdc","promesa de pago"]).sum()
            dur_col  = c_dur
            tmo = grp[dur_col].mean() / 60 if dur_col and pd.api.types.is_numeric_dtype(grp[dur_col]) else 0
            return pd.Series({"Llamadas": llamadas, "Contactos Titular": titular,
                               "Promesas": promesas, "TMO (min)": round(tmo, 1),
                               "Colgadas <30s": 0, "Monto Rec": 0})
        df_asesores = df_gest_real.groupby(c_as).apply(asesor_stats).reset_index()
        df_asesores.rename(columns={c_as: "Asesor"}, inplace=True)
        if df_pago_real is not None and c_asesor and c_monto:
            rec_asesor = df_pago_real.groupby(c_asesor)[c_monto].sum().reset_index()
            rec_asesor.columns = ["Asesor","Monto Rec"]
            df_asesores = df_asesores.merge(rec_asesor, on="Asesor", how="left", suffixes=("","_r"))
            if "Monto Rec_r" in df_asesores.columns:
                df_asesores["Monto Rec"] = df_asesores["Monto Rec_r"].fillna(0)
                df_asesores.drop(columns=["Monto Rec_r"], inplace=True)
        df_asesores["% Contacto"] = (df_asesores["Contactos Titular"] / df_asesores["Llamadas"].clip(1) * 100).round(1)
        df_asesores["% Abandono"] = 0.0
        df_asesores["Conv %"]     = (df_asesores["Promesas"] / df_asesores["Contactos Titular"].clip(1) * 100).round(1)
        df_asesores = df_asesores.sort_values("Monto Rec", ascending=False).reset_index(drop=True)

if df_cart_real is not None:
    c_saldo  = col(df_cart_real, "valor_saldo_deuda","saldo","deuda","monto_deuda","importe")
    c_div    = col(df_cart_real, "division","div","sector")
    c_zona   = col(df_cart_real, "zona","zone","gv","gerencia")
    c_edad   = col(df_cart_real, "edad_consultora","edad","age")
    c_seg    = col(df_cart_real, "segmento_nombre","segmento","segment","categoria")
    c_estado = col(df_cart_real, "direccion_de_residencia_estado","estado","state","entidad")
    if c_saldo:
        META_TOTAL = df_cart_real[c_saldo].sum()
    if c_zona and c_saldo:
        df_gv = (df_cart_real.groupby(c_zona)[c_saldo].agg(["sum","count"])
                 .reset_index()
                 .rename(columns={c_zona:"GV","sum":"Meta","count":"Cuentas"}))
        if df_pago_real is not None:
            c_zona_p = col(df_pago_real, "zona","zone","gv","gerencia")
            c_mp2    = col(df_pago_real, "monto_pagado","valor_pago","importe","monto","pago")
            if c_zona_p and c_mp2:
                rec_gv = df_pago_real.groupby(c_zona_p)[c_mp2].sum().reset_index()
                rec_gv.columns = ["GV","Recuperado"]
                df_gv = df_gv.merge(rec_gv, on="GV", how="left")
                df_gv["Recuperado"] = df_gv["Recuperado"].fillna(0)
            else:
                df_gv["Recuperado"] = df_gv["Meta"] * 0.8
        else:
            df_gv["Recuperado"] = df_gv["Meta"] * 0.8
        df_gv["Cumplimiento %"] = (df_gv["Recuperado"] / df_gv["Meta"].clip(1) * 100).round(1)
        df_gv = df_gv.sort_values("Cumplimiento %", ascending=False).reset_index(drop=True)
    if c_edad and c_saldo:
        df_cart_real["_rango"] = pd.cut(pd.to_numeric(df_cart_real[c_edad], errors="coerce"),
                                         bins=[0,25,35,45,55,65,120],
                                         labels=["18-25","26-35","36-45","46-55","56-65","66+"])
        grp_e = df_cart_real.groupby("_rango", observed=True)
        df_edad = pd.DataFrame({
            "Rango":      grp_e.size().index.astype(str).tolist(),
            "Cuentas":    grp_e.size().values.tolist(),
            "Recuperado": [0]*6,
            "Promesas %": [0]*6,
        })
        if df_pago_real is not None:
            c_edad_p = col(df_pago_real, "edad_consultora","edad")
            c_mp3    = col(df_pago_real, "monto_pagado","valor_pago","importe","monto","pago")
            if c_edad_p and c_mp3:
                df_pago_real["_rango"] = pd.cut(pd.to_numeric(df_pago_real[c_edad_p], errors="coerce"),
                                                  bins=[0,25,35,45,55,65,120],
                                                  labels=["18-25","26-35","36-45","46-55","56-65","66+"])
                rec_e = df_pago_real.groupby("_rango", observed=True)[c_mp3].sum().reset_index()
                rec_e.columns = ["Rango","Recuperado"]
                rec_e["Rango"] = rec_e["Rango"].astype(str)
                df_edad = df_edad.merge(rec_e, on="Rango", how="left", suffixes=("","_r"))
                df_edad["Recuperado"] = df_edad["Recuperado_r"].fillna(0)
                df_edad.drop(columns=["Recuperado_r"], inplace=True)
        df_edad["Ticket Prom"] = (df_edad["Recuperado"] / df_edad["Cuentas"].clip(1)).round(0)
    if c_estado and c_saldo:
        df_estado = (df_cart_real.groupby(c_estado)[c_saldo]
                     .agg(["sum","count"]).reset_index()
                     .rename(columns={c_estado:"Estado","sum":"Meta","count":"Cuentas"})
                     .sort_values("Meta", ascending=False).head(10))
        if df_pago_real is not None:
            c_est_p = col(df_pago_real, "direccion_de_residencia_estado","estado","state","entidad")
            c_mp4   = col(df_pago_real, "monto_pagado","valor_pago","importe","monto","pago")
            if c_est_p and c_mp4:
                rec_est = df_pago_real.groupby(c_est_p)[c_mp4].sum().reset_index()
                rec_est.columns = ["Estado","Recuperado"]
                df_estado = df_estado.merge(rec_est, on="Estado", how="left")
                df_estado["Recuperado"] = df_estado["Recuperado"].fillna(0)
            else:
                df_estado["Recuperado"] = df_estado["Meta"] * 0.8
        else:
            df_estado["Recuperado"] = df_estado["Meta"] * 0.8
        df_estado["Ticket Prom"] = (df_estado["Recuperado"] / df_estado["Cuentas"].clip(1)).round(0)
    if c_seg and c_saldo:
        grp_s = df_cart_real.groupby(c_seg)[c_saldo].agg(["sum","count"]).reset_index()
        grp_s.columns = ["Segmento","Meta","Cuentas"]
        if df_pago_real is not None:
            c_seg_p = col(df_pago_real, "segmento_nombre","segmento","segment","categoria")
            c_mp5   = col(df_pago_real, "monto_pagado","valor_pago","importe","monto","pago")
            if c_seg_p and c_mp5:
                rec_s = df_pago_real.groupby(c_seg_p)[c_mp5].sum().reset_index()
                rec_s.columns = ["Segmento","Recuperado"]
                grp_s = grp_s.merge(rec_s, on="Segmento", how="left")
                grp_s["Recuperado"] = grp_s["Recuperado"].fillna(0)
            else:
                grp_s["Recuperado"] = grp_s["Meta"] * 0.8
        else:
            grp_s["Recuperado"] = grp_s["Meta"] * 0.8
        grp_s["Cumplimiento %"] = (grp_s["Recuperado"] / grp_s["Meta"].clip(1) * 100).round(1)
        df_segmento = grp_s[["Segmento","Cuentas","Recuperado","Cumplimiento %"]]

# ─────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────
col_h1, col_h2, col_h3 = st.columns([3, 1, 1])
with col_h1:
    st.markdown("## 📊 Dashboard Ejecutivo — Cobranza Mayo 2025")
    st.caption("NAtura | Reunión de seguimiento | 09 Jun 2025")
with col_h2:
    pct = RECUPERADO_TOTAL / META_TOTAL * 100
    delta_pct = pct - 100
    st.metric("Recuperación Mayo", f"${RECUPERADO_TOTAL/1e6:.2f}M",
              delta=f"{delta_pct:.1f}% vs meta")
with col_h3:
    st.metric("Contact Rate", f"{CR*100:.1f}%", delta="–3.2pp vs abr")

# ─────────────────────────────────────────────
# TABS
# ─────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
    "1 · Cierre de Mes",
    "2 · Contactabilidad",
    "3 · Indicadores",
    "4 · Operación",
    "5 · Plan de Trabajo",
    "6 · Comparativo Cobranza",
    "7 · Plan de Acción Cobranza",
])

# ══════════════════════════════════════════════
# TAB 1 — CIERRE DE MES
# ══════════════════════════════════════════════
with tab1:
    st.markdown('<div class="sec">Cierre de Mes — Mayo 2025</div>', unsafe_allow_html=True)

    c1,c2,c3,c4,c5 = st.columns(5)
    c1.metric("Asignación por temporalidad", f"${META_TOTAL/1e6:.2f}M")
    c2.metric("Recuperado",           f"${RECUPERADO_TOTAL/1e6:.2f}M",
              delta=f"{(RECUPERADO_TOTAL/META_TOTAL-1)*100:.1f}%")
    c3.metric("Cumplimiento",         f"{RECUPERADO_TOTAL/META_TOTAL*100:.1f}%")
    c4.metric("Promesas generadas",   f"{PROMESAS_GEN:,}")
    c5.metric("Promesas caídas",      f"{PROMESAS_CAIDAS:,}",
              delta=f"{PROMESAS_CAIDAS/PROMESAS_GEN*100:.0f}% caída", delta_color="inverse")

    st.markdown("---")
    col_l, col_r = st.columns([3, 2])

    with col_l:
        st.markdown("**Recuperación acumulada vs Meta**")
        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=df_cierre["Fecha"], y=df_cierre["Meta Acum"],
            name="Meta", line=dict(color=SLATE, dash="dash", width=2),
            fill=None,
        ))
        fig.add_trace(go.Scatter(
            x=df_cierre["Fecha"], y=df_cierre["Recuperado Acum"],
            name="Recuperado", line=dict(color=BLUE, width=3),
            fill="tonexty", fillcolor="rgba(59,130,246,0.08)",
        ))
        fig.add_hrect(y0=META_TOTAL*0.95, y1=META_TOTAL*1.05,
                      fillcolor="rgba(34,197,94,0.05)", line_width=0,
                      annotation_text="±5% meta", annotation_position="top right",
                      annotation_font_color=GREEN)
        apply_layout(fig, height=320, legend=dict(orientation="h", y=1.1))
        fig.update_yaxes(tickprefix="$", tickformat=",.0f")
        st.plotly_chart(fig, use_container_width=True)

    with col_r:
        st.markdown("**Motivos de caída de promesas**")
        df_mot = df_motivos.sort_values("Casos")
        fig2 = go.Figure(go.Bar(
            x=df_mot["Casos"], y=df_mot["Motivo"],
            orientation="h",
            marker_color=[RED if c > 150 else AMBER if c > 80 else SLATE for c in df_mot["Casos"]],
            text=df_mot["Casos"], textposition="outside",
            textfont=dict(color=TEXT),
        ))
        apply_layout(fig2, height=320)
        st.plotly_chart(fig2, use_container_width=True)

    gap = META_TOTAL - RECUPERADO_TOTAL
    st.info(
        f"**Brecha de recuperación: ${gap/1e6:.2f}M** — las **{PROMESAS_CAIDAS} promesas caídas** "
        f"explican ~{PROMESAS_CAIDAS/PROMESAS_GEN*100:.0f}% del desvío. "
        f"Los principales factores: sin liquidez ({df_motivos.iloc[0]['Casos']} casos) "
        f"y contacto perdido ({df_motivos.iloc[1]['Casos']} sin rellamada efectiva)."
    )

# ══════════════════════════════════════════════
# TAB 2 — CONTACTABILIDAD
# ══════════════════════════════════════════════
with tab2:
    st.markdown('<div class="sec">Contactabilidad y Canales Digitales</div>', unsafe_allow_html=True)

    c1,c2,c3,c4,c5 = st.columns(5)
    c1.metric("Total llamadas",        f"{TOTAL_LLAMADAS:,}")
    c2.metric("Contact Rate",         f"{CR*100:.1f}%",  delta="–3.2pp vs abr", delta_color="inverse")
    c3.metric("Titular contactado",   f"{TITULAR:,}")
    c4.metric("Buzón + No contesta",  f"{BUZON+NO_CONT:,}")
    c5.metric("Cuelga al agente",     f"{COLGADO:,}",    delta="Alto", delta_color="inverse")

    st.markdown("---")
    col_l, col_r = st.columns(2)

    with col_l:
        st.markdown("**Contactos y promesas por hora**")
        fig = make_subplots(specs=[[{"secondary_y": True}]])
        fig.add_trace(go.Bar(
            x=df_horario["Hora"], y=df_horario["Contactos"],
            name="Contactos", marker_color=BLUE, opacity=0.8,
        ), secondary_y=False)
        fig.add_trace(go.Scatter(
            x=df_horario["Hora"], y=df_horario["Conv %"],
            name="Conv % promesa", mode="lines+markers",
            line=dict(color=GREEN, width=2), marker=dict(size=6),
        ), secondary_y=True)
        fig.update_layout(**PLOTLY_LAYOUT, height=300,
                          legend=dict(orientation="h", y=1.12))
        fig.update_yaxes(title_text="Contactos", secondary_y=False,
                         gridcolor="#334155")
        fig.update_yaxes(title_text="Conversión %", secondary_y=True,
                         ticksuffix="%", gridcolor="rgba(0,0,0,0)")
        st.plotly_chart(fig, use_container_width=True)
        st.caption("Franja de mayor conversión: **13:00–15:00 h**")

    with col_r:
        st.markdown("**Distribución de intentos de llamada**")
        labels = ["Titular contactado","Buzón","No contesta","Núm inválido","Colgó agente"]
        vals   = [TITULAR, BUZON, NO_CONT, NUM_INV, COLGADO]
        colors = [GREEN, AMBER, SLATE, RED, "#c026d3"]
        fig2 = go.Figure(go.Pie(
            labels=labels, values=vals,
            marker=dict(colors=colors, line=dict(color=DARK, width=2)),
            hole=0.52,
            textinfo="label+percent",
            textfont=dict(color=TEXT),
        ))
        apply_layout(fig2, height=300, showlegend=False)
        st.plotly_chart(fig2, use_container_width=True)

    st.markdown("---")
    st.markdown("**Canales digitales — efectividad para amarrar promesas**")

    cols = st.columns(5)
    canal_icons = {"SMS":"📱","WhatsApp":"💬","Email":"📧","Llamada":"📞","IVR":"🤖"}
    for i, row in df_canal.iterrows():
        with cols[i]:
            icon = canal_icons.get(row["Canal"],"📡")
            color = GREEN if row["Conv prom %"] > 25 else AMBER if row["Conv prom %"] > 18 else RED
            st.markdown(f"""
            <div style="background:{CARD};border-radius:10px;padding:14px;text-align:center;
                        border-top:3px solid {color}">
              <div style="font-size:1.6rem">{icon}</div>
              <div style="font-weight:700;font-size:1rem;color:{TEXT}">{row['Canal']}</div>
              <div style="color:{MUTED};font-size:0.75rem;margin:4px 0">Enviados: {row['Enviados']:,}</div>
              <div style="color:{MUTED};font-size:0.75rem">Resp: {row['Tasa resp %']}%</div>
              <div style="font-size:1.4rem;font-weight:800;color:{color}">{row['Conv prom %']}%</div>
              <div style="color:{MUTED};font-size:0.7rem">conv → promesa</div>
              <div style="margin-top:6px;color:{color};font-weight:700">{row['Promesas']:,} promesas</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.warning(
        "**WhatsApp** lidera en conversión (30.0%). "
        "**SMS** con solo 15.0% de conversión sugiere que el copy no genera urgencia suficiente — "
        "se recomienda A/B test con mensajes personalizados por segmento."
    )

# ══════════════════════════════════════════════
# TAB 3 — INDICADORES
# ══════════════════════════════════════════════
with tab3:
    st.markdown('<div class="sec">Indicadores de Recuperación</div>', unsafe_allow_html=True)
    sub1, sub2, sub3, sub4 = st.tabs(["Por Sector","Por GV","Por Edad","Estado & Segmento"])

    with sub1:
        col_l, col_r = st.columns([3,2])
        with col_l:
            fig = go.Figure()
            fig.add_trace(go.Bar(name="Meta",       x=df_sector["Sector"], y=df_sector["Meta"],
                                 marker_color=SLATE, opacity=0.6))
            fig.add_trace(go.Bar(name="Recuperado", x=df_sector["Sector"], y=df_sector["Recuperado"],
                                 marker_color=[GREEN if p>=80 else AMBER if p>=65 else RED
                                               for p in df_sector["Cumplimiento %"]]))
            apply_layout(fig, height=340, barmode="group",
                         legend=dict(orientation="h", y=1.1),
                         yaxis=dict(tickprefix="$", tickformat=",.0f"))
            st.plotly_chart(fig, use_container_width=True)
        with col_r:
            for _, row in df_sector.sort_values("Cumplimiento %", ascending=False).iterrows():
                c = GREEN if row["Cumplimiento %"]>=80 else AMBER if row["Cumplimiento %"]>=65 else RED
                badge = "badge-green" if row["Cumplimiento %"]>=80 else "badge-amber" if row["Cumplimiento %"]>=65 else "badge-red"
                st.markdown(f"""
                <div style="background:{CARD};border-radius:8px;padding:10px 14px;
                            margin-bottom:6px;display:flex;justify-content:space-between;align-items:center">
                  <span style="color:{TEXT};font-weight:600">{row['Sector']}</span>
                  <span class="badge {badge}">{row['Cumplimiento %']}%</span>
                </div>""", unsafe_allow_html=True)

    with sub2:
        col_l, col_r = st.columns([5,2])
        with col_l:
            fig = go.Figure()
            colors_gv = [GREEN if p>=85 else AMBER if p>=70 else RED for p in df_gv["Cumplimiento %"]]
            fig.add_trace(go.Bar(
                x=df_gv["GV"], y=df_gv["Cumplimiento %"],
                marker_color=colors_gv,
                text=df_gv["Cumplimiento %"].astype(str)+"%",
                textposition="outside", textfont=dict(color=TEXT, size=10),
            ))
            fig.add_hline(y=80, line_dash="dash", line_color=AMBER,
                          annotation_text="Meta 80%", annotation_font_color=AMBER)
            apply_layout(fig, height=350, yaxis=dict(ticksuffix="%", range=[0,115]))
            st.plotly_chart(fig, use_container_width=True)
        with col_r:
            st.markdown(f"**Top 5** <span class='badge badge-green'>Mejor desempeño</span>", unsafe_allow_html=True)
            for _, row in df_gv.head(5).iterrows():
                st.markdown(f"<div style='background:{CARD};border-radius:6px;padding:8px 12px;margin:4px 0;"
                            f"border-left:3px solid {GREEN}'>"
                            f"<b style='color:{TEXT}'>{row['GV']}</b> "
                            f"<span style='color:{GREEN};float:right'>{row['Cumplimiento %']}%</span></div>",
                            unsafe_allow_html=True)
            st.markdown(f"<br>**Bottom 5** <span class='badge badge-red'>Requieren atención</span>", unsafe_allow_html=True)
            for _, row in df_gv.tail(5).iterrows():
                st.markdown(f"<div style='background:{CARD};border-radius:6px;padding:8px 12px;margin:4px 0;"
                            f"border-left:3px solid {RED}'>"
                            f"<b style='color:{TEXT}'>{row['GV']}</b> "
                            f"<span style='color:{RED};float:right'>{row['Cumplimiento %']}%</span></div>",
                            unsafe_allow_html=True)

    with sub3:
        col_l, col_r = st.columns(2)
        with col_l:
            fig = px.bar(df_edad, x="Rango", y="Recuperado",
                         color="Promesas %",
                         color_continuous_scale=["#ef4444","#f59e0b","#22c55e"],
                         text_auto=".2s",
                         labels={"Recuperado":"Recuperado $","Promesas %":"% Con promesa"})
            fig.update_traces(textfont_color=TEXT)
            apply_layout(fig, height=320,
                         yaxis=dict(tickprefix="$", tickformat=",.0f"),
                         coloraxis_colorbar=dict(title="% Promesa", tickfont=dict(color=TEXT)))
            st.plotly_chart(fig, use_container_width=True)
        with col_r:
            fig2 = make_subplots(specs=[[{"secondary_y": True}]])
            fig2.add_trace(go.Bar(
                x=df_edad["Rango"], y=df_edad["Ticket Prom"],
                name="Ticket Prom $", marker_color=PURPLE, opacity=0.8,
            ), secondary_y=False)
            fig2.add_trace(go.Scatter(
                x=df_edad["Rango"], y=df_edad["Promesas %"],
                name="Promesas %", mode="lines+markers",
                line=dict(color=GREEN, width=2), marker=dict(size=7),
            ), secondary_y=True)
            fig2.update_layout(**PLOTLY_LAYOUT, height=320,
                               legend=dict(orientation="h", y=1.12))
            fig2.update_yaxes(tickprefix="$", secondary_y=False, gridcolor="#334155")
            fig2.update_yaxes(ticksuffix="%", secondary_y=True,
                              gridcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig2, use_container_width=True)
        st.info("**Segmento 36-45 años:** mayor recuperación absoluta ($336K) y tasa de promesas más alta (71%). Enfocar refuerzos aquí.")
        st.dataframe(
            df_edad.rename(columns={"Recuperado":"Recuperado $","Ticket Prom":"Ticket $"})
                   .style.format({"Recuperado $":"${:,.0f}","Ticket $":"${:,.0f}","Promesas %":"{:.0f}%"}),
            use_container_width=True, hide_index=True,
        )

    with sub4:
        col_l, col_r = st.columns(2)
        with col_l:
            st.markdown("**Recuperación por Estado (Top 10)**")
            fig = px.bar(df_estado.sort_values("Recuperado"),
                         x="Recuperado", y="Estado", orientation="h",
                         color="Recuperado",
                         color_continuous_scale=["#1e3a5f","#3b82f6","#22c55e"],
                         text_auto=".2s")
            fig.update_traces(textfont_color=TEXT)
            apply_layout(fig, height=360,
                         xaxis=dict(tickprefix="$", tickformat=",.0f"),
                         coloraxis_showscale=False)
            st.plotly_chart(fig, use_container_width=True)
        with col_r:
            st.markdown("**Recuperación y cumplimiento por Segmento**")
            fig2 = make_subplots(specs=[[{"secondary_y": True}]])
            fig2.add_trace(go.Bar(
                x=df_segmento["Segmento"], y=df_segmento["Recuperado"],
                name="Recuperado $", marker_color=BLUE, opacity=0.8,
            ), secondary_y=False)
            fig2.add_trace(go.Scatter(
                x=df_segmento["Segmento"], y=df_segmento["Cumplimiento %"],
                name="% Cumplimiento", mode="lines+markers",
                line=dict(color=AMBER, width=2), marker=dict(size=8),
            ), secondary_y=True)
            fig2.update_layout(**PLOTLY_LAYOUT, height=360,
                               legend=dict(orientation="h", y=1.12))
            fig2.update_yaxes(tickprefix="$", tickformat=",.0f",
                              secondary_y=False, gridcolor="#334155")
            fig2.update_yaxes(ticksuffix="%", secondary_y=True,
                              gridcolor="rgba(0,0,0,0)", range=[0,110])
            st.plotly_chart(fig2, use_container_width=True)
        st.markdown("**Detalle por segmento**")
        st.dataframe(
            df_segmento.style.format({"Recuperado":"${:,.0f}","Cumplimiento %":"{:.0f}%"}),
            use_container_width=True, hide_index=True,
        )

# ══════════════════════════════════════════════
# TAB 4 — OPERACIÓN
# ══════════════════════════════════════════════
with tab4:
    st.markdown('<div class="sec">Operación de Asesores — Mayo 2025</div>', unsafe_allow_html=True)

    avg_tmo     = df_asesores["TMO (min)"].mean()
    total_llam  = TOTAL_LLAMADAS
    avg_abandono= df_asesores["% Abandono"].mean()
    avg_conv    = df_asesores["Conv %"].mean()

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Total llamadas",       f"{total_llam:,}")
    c2.metric("TMO promedio",         f"{avg_tmo:.1f} min")
    c3.metric("% Abandono <30s",      f"{avg_abandono:.1f}%",
              delta="Revisar apertura", delta_color="inverse")
    c4.metric("Conv. Contacto→Promesa", f"{avg_conv:.1f}%")

    st.markdown("---")
    col_l, col_r = st.columns([3,2])

    with col_l:
        st.markdown("**Ranking de asesores — monto recuperado y conversión**")
        fig = make_subplots(specs=[[{"secondary_y": True}]])
        fig.add_trace(go.Bar(
            x=df_asesores["Asesor"], y=df_asesores["Monto Rec"],
            name="Monto Rec $", marker_color=BLUE, opacity=0.85,
        ), secondary_y=False)
        fig.add_trace(go.Scatter(
            x=df_asesores["Asesor"], y=df_asesores["Conv %"],
            name="Conv %", mode="lines+markers",
            line=dict(color=GREEN, width=2), marker=dict(size=7),
        ), secondary_y=True)
        fig.update_layout(**PLOTLY_LAYOUT, height=320,
                          legend=dict(orientation="h", y=1.12))
        fig.update_yaxes(tickprefix="$", tickformat=",.0f",
                         secondary_y=False, gridcolor="#334155")
        fig.update_yaxes(ticksuffix="%", secondary_y=True,
                         gridcolor="rgba(0,0,0,0)")
        st.plotly_chart(fig, use_container_width=True)

    with col_r:
        st.markdown("**TMO y % Abandono por asesor**")
        fig2 = go.Figure()
        fig2.add_trace(go.Bar(
            name="TMO (min)",    x=df_asesores["Asesor"], y=df_asesores["TMO (min)"],
            marker_color=[RED if t>6 else AMBER if t>4.5 else GREEN for t in df_asesores["TMO (min)"]],
        ))
        fig2.add_trace(go.Scatter(
            name="% Abandono",   x=df_asesores["Asesor"], y=df_asesores["% Abandono"],
            mode="lines+markers", line=dict(color=RED, width=2, dash="dot"),
            marker=dict(size=7), yaxis="y2",
        ))
        fig2.update_layout(
            **PLOTLY_LAYOUT, height=320, barmode="group",
            legend=dict(orientation="h", y=1.12),
            yaxis2=dict(overlaying="y", side="right", ticksuffix="%",
                        gridcolor="rgba(0,0,0,0)", showgrid=False),
        )
        st.plotly_chart(fig2, use_container_width=True)

    st.markdown("---")
    col_l2, col_r2 = st.columns(2)

    with col_l2:
        st.markdown("**Frecuencia de objeciones difíciles**")
        fig3 = go.Figure(go.Bar(
            x=df_objeciones["Frecuencia"],
            y=df_objeciones["Objeción"],
            orientation="h",
            marker_color=[RED if f>2000 else AMBER if f>1000 else SLATE
                          for f in df_objeciones["Frecuencia"]],
            text=df_objeciones["Frecuencia"], textposition="outside",
            textfont=dict(color=TEXT),
        ))
        apply_layout(fig3, height=280)
        st.plotly_chart(fig3, use_container_width=True)

    with col_r2:
        st.markdown("**Tasa de resolución exitosa por objeción**")
        fig4 = go.Figure(go.Bar(
            x=df_objeciones["Resolución exitosa %"],
            y=df_objeciones["Objeción"],
            orientation="h",
            marker_color=[GREEN if p>=65 else AMBER if p>=50 else RED
                          for p in df_objeciones["Resolución exitosa %"]],
            text=df_objeciones["Resolución exitosa %"].astype(str)+"%",
            textposition="outside", textfont=dict(color=TEXT),
        ))
        apply_layout(fig4, height=280, xaxis=dict(ticksuffix="%", range=[0,100]))
        st.plotly_chart(fig4, use_container_width=True)

    st.warning(
        f"**Puntos críticos:** "
        f"'No tengo dinero' es la objeción más frecuente (2,840 casos) con solo **38%** de resolución — "
        f"el script actual no es efectivo. "
        f"% Abandono promedio de **{avg_abandono:.1f}%**: los clientes cuelgan antes de que el asesor pueda hablar. "
        f"Revisar la apertura de llamada y los primeros 10 segundos del guión."
    )

    st.markdown("---")
    st.markdown("**Detalle completo por asesor**")
    display_df = df_asesores[["Asesor","Llamadas","TMO (min)","% Contacto","Conv %","% Abandono","Monto Rec"]].copy()
    display_df["Monto Rec"] = display_df["Monto Rec"].apply(lambda x: f"${x:,.0f}")
    st.dataframe(display_df, use_container_width=True, hide_index=True)

# ══════════════════════════════════════════════
# TAB 5 — PLAN DE TRABAJO
# ══════════════════════════════════════════════
with tab5:
    st.markdown('<div class="sec">Plan de Trabajo — Junio 2025</div>', unsafe_allow_html=True)

    gap_jun = META_JUN - PROY_JUN
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Meta Junio",           f"${META_JUN/1e6:.1f}M")
    c2.metric("Proyección base",      f"${PROY_JUN/1e6:.2f}M",
              delta=f"{(PROY_JUN/META_JUN-1)*100:.1f}%", delta_color="inverse")
    c3.metric("Gap a cerrar",         f"${gap_jun/1e6:.2f}M", delta_color="inverse")
    c4.metric("Pipeline de promesas", f"${PIPELINE_PROM/1e6:.2f}M")

    st.markdown("---")
    col_l, col_r = st.columns(2)

    with col_l:
        st.markdown("**Proyección acumulada Junio — 3 escenarios**")
        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=df_proy["Fecha"], y=df_proy["Meta"],
            name="Meta", line=dict(color=SLATE, dash="dash", width=2),
        ))
        fig.add_trace(go.Scatter(
            x=df_proy["Fecha"], y=df_proy["Optimista"],
            name="Optimista", line=dict(color=GREEN, width=2),
            fill="tonexty", fillcolor="rgba(34,197,94,0.05)",
        ))
        fig.add_trace(go.Scatter(
            x=df_proy["Fecha"], y=df_proy["Base"],
            name="Base", line=dict(color=BLUE, width=3),
        ))
        fig.add_trace(go.Scatter(
            x=df_proy["Fecha"], y=df_proy["Pesimista"],
            name="Pesimista", line=dict(color=RED, width=2, dash="dot"),
            fill="tonexty", fillcolor="rgba(239,68,68,0.05)",
        ))
        apply_layout(fig, height=360, legend=dict(orientation="h", y=1.1))
        fig.update_yaxes(tickprefix="$", tickformat=",.0f")
        st.plotly_chart(fig, use_container_width=True)

    with col_r:
        st.markdown("**Impacto estimado por palanca de acción**")
        df_p = df_palancas.sort_values("Impacto estimado $")
        fig2 = go.Figure(go.Bar(
            x=df_p["Impacto estimado $"], y=df_p["Palanca"],
            orientation="h", marker_color=BLUE,
            text=df_p["Impacto estimado $"].apply(lambda x: f"${x:,.0f}"),
            textposition="outside", textfont=dict(color=TEXT),
        ))
        total_impacto = df_palancas["Impacto estimado $"].sum()
        fig2.add_vline(x=gap_jun, line_dash="dash", line_color=RED,
                       annotation_text=f"Gap: ${gap_jun:,.0f}",
                       annotation_font_color=RED)
        apply_layout(fig2, height=360, xaxis=dict(tickprefix="$", tickformat=",.0f"))
        st.plotly_chart(fig2, use_container_width=True)
        st.caption(f"Impacto total estimado: **${total_impacto:,.0f}** — cubre "
                   f"**{total_impacto/gap_jun*100:.0f}%** del gap de Junio")

    st.markdown("---")
    st.markdown("**Compromisos del equipo**")
    for _, row in df_acciones.iterrows():
        color = RED if row["Plazo"] == "Inmediato" else AMBER
        badge_cls = "badge-red" if row["Plazo"] == "Inmediato" else "badge-amber"
        st.markdown(
            f"<div style='background:{CARD};border-radius:8px;padding:10px 16px;margin:5px 0;"
            f"display:flex;align-items:center;gap:12px'>"
            f"<span class='badge {badge_cls}'>{row['Plazo']}</span>"
            f"<span style='color:{MUTED};min-width:130px;font-size:0.85rem'>{row['Área']}</span>"
            f"<span style='color:{TEXT}'>{row['Acción']}</span>"
            f"</div>",
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)
    st.success(
        f"**Resumen ejecutivo:** Con las palancas propuestas se puede cerrar hasta **${total_impacto/1e6:.2f}M** "
        f"adicional, cubriendo el **{total_impacto/gap_jun*100:.0f}%** del gap. "
        f"La prioridad #1 esta semana es activar la rellamada a las **{PROMESAS_CAIDAS} promesas caídas** de Mayo "
        f"y lanzar el A/B test de WhatsApp antes del miércoles."
    )

# ══════════════════════════════════════════════
# TAB 6 — COMPARATIVO COBRANZA
# ══════════════════════════════════════════════
with tab6:
    st.markdown('<div class="sec">Comparativo Cobranza por Tramo</div>', unsafe_allow_html=True)
    if df_comparativo is None:
        st.markdown(
            f"<div style='text-align:center;padding:50px 20px;background:{CARD};border-radius:16px;"
            f"border:2px dashed #e2e8f0'>"
            f"<div style='font-size:2.5rem'>📑</div>"
            f"<div style='font-size:1.2rem;font-weight:700;color:{TEXT};margin:10px 0'>Sube el archivo Comparativo</div>"
            f"<div style='color:{MUTED};font-size:0.9rem'>Carga el Excel con las hojas <b>'TABLA DE PAGOS'</b> y <b>'COMPARATIVO'</b> "
            f"en la pantalla de inicio (📑 Comparativo Cobranza) y presiona el botón para ver el dashboard.<br>"
            f"El sistema sincronizará automáticamente la Cobranza $ por tramo desde los totales de Pagos.</div>"
            f"</div>", unsafe_allow_html=True)
        if DEBUG_COMP is not None:
            st.markdown("---")
            with st.expander("🔍 Diagnóstico del archivo subido (clic para ver detalle)"):
                st.write("**Hojas encontradas:**", DEBUG_COMP.get("hojas"))
                st.write("**Hoja de Pagos detectada:**", DEBUG_COMP.get("hoja_pagos"))
                st.write("**Hoja de Comparativo detectada:**", DEBUG_COMP.get("hoja_comparativo"))
                st.write("**Fila de encabezados T1..Tn (Pagos):**", DEBUG_COMP.get("header_row_pagos"))
                st.write("**Columnas de tramo detectadas (Pagos):**", DEBUG_COMP.get("tramo_cols"))
                st.write("**Fila TOTAL (Pagos):**", DEBUG_COMP.get("total_row_pagos"))
                st.write("**Totales por tramo (fila TOTAL):**", DEBUG_COMP.get("totales_tramo"))
                st.write("**Fila de encabezados (Comparativo):**", DEBUG_COMP.get("header_row_comparativo"))
                st.write("**Columnas detectadas (Comparativo):**", DEBUG_COMP.get("cols_comparativo"))
                st.write("**Valores de la columna Tramo (Comparativo):**", DEBUG_COMP.get("tramo_vals_comparativo"))
                if "error" in DEBUG_COMP:
                    st.error(DEBUG_COMP["error"])
    else:
        r = RESUMEN_COMP
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Cartera asignada", f"${r['total_cartera']/1e6:.2f}M")
        c2.metric("Objetivo del mes", f"${r['total_objetivo']/1e6:.2f}M")
        c3.metric("Cobranza actual", f"${r['total_cobranza']/1e6:.2f}M",
                  delta=f"{r['cumplimiento']-100:.1f}pp vs objetivo")
        c4.metric("Días pendientes", f"{r['dias_pendientes']:.0f} de {r['dias_totales_mes']}")
        st.markdown("---")

        cl, cr = st.columns(2)
        with cl:
            st.markdown("**Cartera, Objetivo y Cobranza por Tramo**")
            fig = go.Figure()
            fig.add_trace(go.Bar(name="Cartera $", x=df_comparativo["Tramo"], y=df_comparativo["Cartera $"], marker_color="#cbd5e1", opacity=0.85))
            fig.add_trace(go.Bar(name="Objetivo $", x=df_comparativo["Tramo"], y=df_comparativo["Objetivo $"], marker_color=SLATE, opacity=0.85))
            fig.add_trace(go.Bar(name="Cobranza $", x=df_comparativo["Tramo"], y=df_comparativo["Cobranza $"], marker_color=BLUE))
            apply_layout(fig, height=340, barmode="group", legend=dict(orientation="h", y=1.1),
                         yaxis=dict(tickprefix="$", tickformat=",.0f"))
            st.plotly_chart(fig, use_container_width=True)
        with cr:
            st.markdown("**% de Cumplimiento de Objetivo por Tramo**")
            df_cs = df_comparativo.sort_values("Cumplimiento %")
            fig2 = go.Figure(go.Bar(x=df_cs["Cumplimiento %"], y=df_cs["Tramo"], orientation="h",
                marker_color=[GREEN if p>=100 else AMBER if p>=70 else RED for p in df_cs["Cumplimiento %"]],
                text=df_cs["Cumplimiento %"].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "—"),
                textposition="outside", textfont=dict(color=TEXT)))
            fig2.add_vline(x=100, line_dash="dash", line_color=GREEN, annotation_text="Meta 100%", annotation_font_color=GREEN)
            apply_layout(fig2, height=340, xaxis=dict(ticksuffix="%"))
            st.plotly_chart(fig2, use_container_width=True)

        if df_pagos_diario is not None and len(df_pagos_diario) > 0:
            st.markdown("---")
            st.markdown("**Cobranza diaria y acumulado**")
            fig3 = make_subplots(specs=[[{"secondary_y": True}]])
            fig3.add_trace(go.Bar(x=df_pagos_diario["Fecha"], y=df_pagos_diario["Total Día"], name="Cobranza diaria",
                                  marker_color=BLUE, opacity=0.75), secondary_y=False)
            fig3.add_trace(go.Scatter(x=df_pagos_diario["Fecha"], y=df_pagos_diario["Acumulado"], name="Acumulado",
                                      mode="lines+markers", line=dict(color=GREEN, width=3), marker=dict(size=5)), secondary_y=True)
            fig3.add_hline(y=r["promedio_diario"], line_dash="dash", line_color=AMBER,
                           annotation_text=f"Promedio diario: ${r['promedio_diario']:,.0f}", annotation_font_color=AMBER)
            fig3.update_layout(**PLOTLY_LAYOUT, height=340, legend=dict(orientation="h", y=1.1))
            fig3.update_yaxes(tickprefix="$", tickformat=",.0f", secondary_y=False, gridcolor="#e2e8f0")
            fig3.update_yaxes(tickprefix="$", tickformat=",.0f", secondary_y=True, gridcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig3, use_container_width=True)

        st.markdown("---")
        st.markdown("**Detalle por Tramo (sincronizado desde Pagos)**")
        fmt = {"Cartera $": "${:,.0f}", "Objetivo $": "${:,.0f}", "Cobranza $": "${:,.0f}",
               "Cumplimiento %": "{:.1f}%"}
        if "Objetivo %" in df_comparativo.columns:
            fmt["Objetivo %"] = "{:.1f}%"
        if "Cuentas" in df_comparativo.columns:
            fmt["Cuentas"] = "{:,.0f}"
        st.dataframe(df_comparativo.style.format(fmt), use_container_width=True, hide_index=True)

        st.markdown("---")
        if r["proyeccion_total"] >= r["total_objetivo"]:
            st.success(f"**Proyección de cierre de mes:** ${r['proyeccion_total']/1e6:.2f}M — al ritmo actual "
                       f"(${r['promedio_diario']:,.0f}/día durante {r['dias_pendientes']:.0f} días restantes) "
                       f"se **superaría el objetivo** de ${r['total_objetivo']/1e6:.2f}M "
                       f"(cumplimiento actual {r['cumplimiento']:.1f}%).")
        else:
            falta = r["total_objetivo"] - r["proyeccion_total"]
            st.warning(f"**Proyección de cierre de mes:** ${r['proyeccion_total']/1e6:.2f}M — al ritmo actual "
                       f"(${r['promedio_diario']:,.0f}/día durante {r['dias_pendientes']:.0f} días restantes) "
                       f"**faltarían ${falta/1e6:.2f}M** para alcanzar el objetivo de ${r['total_objetivo']/1e6:.2f}M "
                       f"(cumplimiento actual {r['cumplimiento']:.1f}%).")

# ══════════════════════════════════════════════
# TAB 7 — PLAN DE ACCIÓN COBRANZA (módulo independiente)
# ══════════════════════════════════════════════
with tab7:
    st.markdown('<div class="sec">📋 Plan de Acción Cobranza</div>', unsafe_allow_html=True)

    if "pac_df" not in st.session_state:
        st.session_state.pac_df = None
        st.session_state.pac_filename = None
        st.session_state.pac_loaded_at = None
        st.session_state.pac_history = []  # lista de snapshots: {fecha, df, cols, filename}

    # ── ESTADO A: sin archivo cargado ──
    if st.session_state.pac_df is None:
        st.markdown(
            f"<div style='text-align:center;padding:40px 20px;background:{CARD};border-radius:16px;"
            f"border:2px dashed #e2e8f0'>"
            f"<div style='font-size:2.5rem'>📂</div>"
            f"<div style='font-size:1.2rem;font-weight:700;color:{PAC_NAVY};margin:10px 0'>"
            f"Sube el archivo de cartera para generar el Plan de Acción</div>"
            f"<div style='color:#64748b;font-size:0.9rem;max-width:680px;margin:0 auto'>"
            f"El archivo .xlsx debe contener las columnas: <b>id/folio</b>, <b>aging_de_morosidad</b>, "
            f"<b>dias_de_morosidad</b>, <b>saldo_insoluto</b>, <b>valor_original_deuda</b>, <b>pago_actual</b> "
            f"(se toma de la columna <b>{PAC_PAGO_LETRA}</b> del Excel), "
            f"<b>status_rep</b>, <b>edad_consultora</b>, <b>segmentacion_rep</b>, <b>division</b>, <b>zona</b>, "
            f"<b>campana_numero</b>, <b>score_riesgo</b>, <b>numero_telefono_celular</b>, <b>correo_electronico</b>."
            f"</div></div>", unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        fecha_corte = st.date_input("Fecha de corte de este archivo", value=pd.Timestamp.now().date(), key="pac_fecha_corte")
        f_pac = st.file_uploader("Selecciona archivo .xlsx", type=["xlsx"], key="pac_uploader")
        if f_pac is not None:
            try:
                df_pac_raw = pd.read_excel(f_pac, engine="openpyxl")
            except Exception as e:
                st.error(f"No se pudo leer el archivo: {e}")
                df_pac_raw = None
            if df_pac_raw is not None:
                encontradas, faltantes = pac_validar_columnas(df_pac_raw)
                if faltantes:
                    st.error("❌ Faltan columnas requeridas: " + ", ".join(faltantes))
                else:
                    df_proc = pac_procesar(df_pac_raw, encontradas)
                    st.session_state.pac_df = df_proc
                    st.session_state.pac_cols = encontradas
                    st.session_state.pac_filename = f_pac.name
                    st.session_state.pac_loaded_at = pd.Timestamp.now()
                    historial = [h for h in st.session_state.pac_history if h["fecha"] != fecha_corte]
                    historial.append({"fecha": fecha_corte, "df": df_proc, "cols": encontradas, "filename": f_pac.name})
                    st.session_state.pac_history = sorted(historial, key=lambda h: h["fecha"])
                    st.rerun()
        if st.session_state.pac_history:
            st.markdown("<br>", unsafe_allow_html=True)
            st.caption(f"Cortes guardados en esta sesión: {', '.join(h['fecha'].strftime('%d/%m') for h in st.session_state.pac_history)}")
            if st.button("🗑 Borrar historial de cortes"):
                st.session_state.pac_history = []
                st.rerun()

    # ── ESTADO B: archivo cargado y procesado ──
    else:
        df_pac = st.session_state.pac_df
        c = st.session_state.pac_cols

        col_h1, col_h2, col_h3 = st.columns([3, 1, 1])
        with col_h1:
            st.markdown(
                f"**📄 {st.session_state.pac_filename}** &nbsp;|&nbsp; "
                f"Cargado: {st.session_state.pac_loaded_at.strftime('%d/%m/%Y %H:%M')} "
                f"&nbsp;|&nbsp; Cortes en historial: {len(st.session_state.pac_history)}"
            )
        with col_h2:
            html_report = pac_export_html(df_pac, c)
            st.download_button(
                "⬇ Descargar reporte HTML", data=html_report,
                file_name=f"plan_accion_natura_{pd.Timestamp.now().strftime('%Y%m%d')}.html",
                mime="text/html", use_container_width=True,
            )
        with col_h3:
            if st.button("🔄 Cargar nuevo corte", use_container_width=True):
                st.session_state.pac_df = None
                st.session_state.pac_filename = None
                st.session_state.pac_loaded_at = None
                st.rerun()

        st.markdown("---")

        orden_t = ["T1","T2","T3","T4","T5","T6","T7"]
        orden_r = ["Crítico","Alto","Preventivo","Estable"]
        orden_edad = ["18–30","31–45","46–60","61+"]

        pac1, pac2, pac3, pac4, pac5, pac6, pac7 = st.tabs([
            "1 · Resumen ejecutivo","2 · Matriz de riesgo","3 · Por temporalidad",
            "4 · Plan operativo","5 · Reglas","6 · Zonas & División","7 · Seguimiento diario",
        ])

        # ── PAC TAB 1 — RESUMEN EJECUTIVO ──
        with pac1:
            total_consultoras = len(df_pac)
            saldo_total  = df_pac[c["saldo_insoluto"]].sum()
            deuda_total  = df_pac[c["valor_original_deuda"]].sum()
            pagos_total  = df_pac[c["pago_actual"]].sum()
            pct_recup    = (pagos_total/deuda_total*100) if deuda_total else 0.0
            criticas     = df_pac[df_pac["RiesgoMigracion"]=="Crítico"]

            k1,k2,k3,k4,k5,k6 = st.columns(6)
            k1.metric("Consultoras en cartera", f"{total_consultoras:,}")
            k2.metric("Deuda asignada", f"${saldo_total/1e6:.2f}M")
            k3.metric("Deuda original total", f"${deuda_total/1e6:.2f}M")
            k4.metric("Pagos registrados", f"${pagos_total/1e6:.2f}M")
            k5.metric("% de recuperación general", f"{pct_recup:.1f}%")
            k6.metric("Migran ≤7 días", f"{len(criticas):,}", delta=f"${criticas[c['saldo_insoluto']].sum()/1e6:.2f}M en riesgo")

            st.markdown("---")
            cl, cm, cr = st.columns(3)
            with cl:
                st.markdown("**Saldo por Temporalidad**")
                saldo_por_t = df_pac.groupby("Temporalidad")[c["saldo_insoluto"]].sum().reindex(orden_t).fillna(0)
                fig = go.Figure(go.Bar(x=orden_t, y=saldo_por_t.values, marker_color=PAC_TEAL))
                apply_layout(fig, height=300, yaxis=dict(tickprefix="$", tickformat=",.0f"))
                st.plotly_chart(fig, use_container_width=True)
            with cm:
                st.markdown("**Riesgo de Migración**")
                riesgo_dist = df_pac["RiesgoMigracion"].value_counts().reindex(orden_r).fillna(0)
                fig2 = go.Figure(go.Pie(labels=orden_r, values=riesgo_dist.values, hole=0.55,
                    marker_colors=[PAC_RED,PAC_CORAL,PAC_AMBER,PAC_TEAL2]))
                apply_layout(fig2, height=300)
                st.plotly_chart(fig2, use_container_width=True)
            with cr:
                st.markdown("**Consultoras por Temporalidad**")
                cuentas_por_t = df_pac["Temporalidad"].value_counts().reindex(orden_t).fillna(0)
                fig3 = go.Figure(go.Bar(x=orden_t, y=cuentas_por_t.values, marker_color=PAC_NAVY))
                apply_layout(fig3, height=300)
                st.plotly_chart(fig3, use_container_width=True)

        # ── PAC TAB 2 — MATRIZ DE RIESGO ──
        with pac2:
            canales = {
                "Crítico":   "Llamada diaria + Llamada de seguimiento + SMS cada 48h + Reminder diario",
                "Alto":      "Llamada 3× semana + Llamada de seguimiento + SMS 2× semana",
                "Preventivo":"Llamada 2× semana + SMS semanal + Reminder semanal",
            }
            cc1, cc2, cc3 = st.columns(3)
            for col_obj, nivel, color in [(cc1,"Crítico",PAC_RED),(cc2,"Alto",PAC_CORAL),(cc3,"Preventivo",PAC_AMBER)]:
                sub = df_pac[df_pac["RiesgoMigracion"]==nivel]
                with col_obj:
                    st.markdown(
                        f"<div style='background:{CARD};border-radius:10px;padding:16px;border-left:4px solid {color}'>"
                        f"<div style='font-weight:700;color:{color};font-size:1.1rem'>{nivel}</div>"
                        f"<div style='font-size:1.6rem;font-weight:800;color:{PAC_NAVY}'>{len(sub):,} consultoras</div>"
                        f"<div style='color:#64748b'>Saldo: ${sub[c['saldo_insoluto']].sum():,.0f}</div>"
                        f"<div style='margin-top:8px;font-size:0.85rem;color:{PAC_NAVY}'><b>Canales:</b> {canales[nivel]}</div>"
                        f"</div>", unsafe_allow_html=True)

            st.markdown("---")
            st.markdown("**Riesgo × Rango de Edad**")
            tab_re = df_pac[df_pac["RiesgoMigracion"].isin(["Crítico","Alto","Preventivo"])]
            pvt = tab_re.pivot_table(index="RangoEdad", columns="RiesgoMigracion", aggfunc="size", fill_value=0)
            pvt = pvt.reindex(orden_edad).fillna(0)
            fig4 = go.Figure()
            for nivel, color in [("Crítico",PAC_RED),("Alto",PAC_CORAL),("Preventivo",PAC_AMBER)]:
                if nivel in pvt.columns:
                    fig4.add_trace(go.Bar(name=nivel, x=pvt.index, y=pvt[nivel], marker_color=color))
            apply_layout(fig4, height=320, barmode="stack", legend=dict(orientation="h", y=1.15))
            st.plotly_chart(fig4, use_container_width=True)

            col_estado_geo = pac_col_por_letra(df_pac, PAC_ESTADO_GEO_LETRA)
            if col_estado_geo is not None:
                st.markdown("---")
                st.markdown("**Riesgo por Estado**")
                tab_estado = df_pac[df_pac["RiesgoMigracion"].isin(["Crítico","Alto","Preventivo"])]
                pvt_estado = tab_estado.pivot_table(index=col_estado_geo, columns="RiesgoMigracion", aggfunc="size", fill_value=0)
                pvt_estado["Total"] = pvt_estado.sum(axis=1)
                pvt_estado = pvt_estado.sort_values("Total", ascending=False).drop(columns="Total").head(15)
                fig_estado = go.Figure()
                for nivel, color in [("Crítico",PAC_RED),("Alto",PAC_CORAL),("Preventivo",PAC_AMBER)]:
                    if nivel in pvt_estado.columns:
                        fig_estado.add_trace(go.Bar(name=nivel, x=pvt_estado.index, y=pvt_estado[nivel], marker_color=color))
                apply_layout(fig_estado, height=360, barmode="stack", legend=dict(orientation="h", y=1.15))
                st.plotly_chart(fig_estado, use_container_width=True)

        # ── PAC TAB 3 — POR TEMPORALIDAD ──
        with pac3:
            filas = []
            for t in orden_t:
                sub = df_pac[df_pac["Temporalidad"]==t]
                if len(sub)==0:
                    filas.append([t,0,0.0,0.0,0.0,0.0,"—"])
                    continue
                riesgo_dom = sub["RiesgoMigracion"].value_counts().idxmax() if len(sub) else "—"
                filas.append([
                    t, len(sub), len(sub)/len(df_pac)*100,
                    sub[c["dias_de_morosidad"]].mean(),
                    sub[c["saldo_insoluto"]].sum(),
                    sub[c["saldo_insoluto"]].sum()/df_pac[c["saldo_insoluto"]].sum()*100 if df_pac[c["saldo_insoluto"]].sum() else 0,
                    riesgo_dom,
                ])
            tabla_t = pd.DataFrame(filas, columns=["Temporalidad","Consultoras","% del total","Días prom. mora",
                                                     "Valor saldo","% del saldo","Riesgo dominante"])
            st.dataframe(tabla_t.style.format({
                "% del total":"{:.1f}%","Días prom. mora":"{:.0f}","Valor saldo":"${:,.0f}","% del saldo":"{:.1f}%",
            }), use_container_width=True, hide_index=True)

            st.markdown("---")
            colA, colB = st.columns(2)
            with colA:
                st.markdown("**Consultoras por Temporalidad**")
                fig5 = go.Figure(go.Bar(x=tabla_t["Temporalidad"], y=tabla_t["Consultoras"], marker_color=PAC_TEAL))
                apply_layout(fig5, height=300)
                st.plotly_chart(fig5, use_container_width=True)
            with colB:
                st.markdown("**Saldo por Temporalidad**")
                fig6 = go.Figure(go.Bar(x=tabla_t["Temporalidad"], y=tabla_t["Valor saldo"], marker_color=PAC_CORAL))
                apply_layout(fig6, height=300, yaxis=dict(tickprefix="$", tickformat=",.0f"))
                st.plotly_chart(fig6, use_container_width=True)

            st.markdown("---")
            st.markdown("**Segmentación**")
            seg_tab = df_pac.groupby(c["segmentacion_rep"]).agg(
                Consultoras=(c["id"] if "id" in c else c["saldo_insoluto"], "count"),
                Saldo=(c["saldo_insoluto"], "sum"),
            ).reset_index().rename(columns={c["segmentacion_rep"]:"Segmentación"}).sort_values("Saldo", ascending=False)
            st.dataframe(seg_tab.style.format({"Saldo":"${:,.0f}"}), use_container_width=True, hide_index=True)

        # ── PAC TAB 4 — PLAN OPERATIVO ──
        with pac4:
            n_crit = len(df_pac[df_pac["RiesgoMigracion"]=="Crítico"])
            n_alto = len(df_pac[df_pac["RiesgoMigracion"]=="Alto"])
            n_prev = len(df_pac[df_pac["RiesgoMigracion"]=="Preventivo"])

            semana = st.radio("Semana", [
                "Semana 1 · Identificación y prevención",
                "Semana 2 · Conversión de promesas",
                "Semana 3 · Contención de migración",
                "Semana 4 · Cierre de mes",
            ], horizontal=True)

            planes = {
                "Semana 1 · Identificación y prevención": [
                    ("Lunes",   f"Identificar {n_prev:,} consultoras en riesgo Preventivo y armar lista de llamadas"),
                    ("Martes",  f"Llamadas a {n_prev:,} consultoras preventivas"),
                    ("Miércoles", f"SMS de recordatorio a {n_prev:,} consultoras preventivas"),
                    ("Jueves",  f"Llamadas de seguimiento a {n_alto:,} consultoras de riesgo Alto"),
                    ("Viernes", f"Revisión de promesas de pago generadas en la semana"),
                    ("Sábado",  f"Cierre semanal y reporte de avance a supervisión"),
                ],
                "Semana 2 · Conversión de promesas": [
                    ("Lunes",   f"Llamadas a {n_alto:,} consultoras de riesgo Alto para confirmar promesas"),
                    ("Martes",  f"SMS de confirmación de pago a consultoras con promesa activa"),
                    ("Miércoles", f"Llamadas de seguimiento a promesas no cumplidas"),
                    ("Jueves",  f"Reminder a {n_prev:,} consultoras preventivas"),
                    ("Viernes", f"Llamadas a {n_crit:,} consultoras críticas"),
                    ("Sábado",  f"Reporte de conversión de promesas de la semana"),
                ],
                "Semana 3 · Contención de migración": [
                    ("Lunes",   f"Llamada diaria a {n_crit:,} consultoras críticas (≤7 días para migrar)"),
                    ("Martes",  f"SMS cada 48h a {n_crit:,} consultoras críticas"),
                    ("Miércoles", f"Llamadas 3× semana a {n_alto:,} consultoras de riesgo Alto"),
                    ("Jueves",  f"Reminder diario a consultoras críticas"),
                    ("Viernes", f"Llamada de seguimiento a consultoras críticas y altas"),
                    ("Sábado",  f"Reporte de contención: consultoras que evitaron migrar"),
                ],
                "Semana 4 · Cierre de mes": [
                    ("Lunes",   f"Llamadas finales a {n_crit:,} consultoras críticas pendientes"),
                    ("Martes",  f"SMS de último recordatorio a consultoras con saldo pendiente"),
                    ("Miércoles", f"Llamadas de cierre a {n_alto:,} consultoras de riesgo Alto"),
                    ("Jueves",  f"Consolidado de pagos registrados vs meta del mes"),
                    ("Viernes", f"Llamadas de cierre a consultoras preventivas restantes"),
                    ("Sábado",  f"Reporte ejecutivo de cierre de mes para gerencia"),
                ],
            }
            for dia, accion in planes[semana]:
                st.markdown(
                    f"<div style='background:{CARD};border-radius:8px;padding:10px 16px;margin:5px 0;"
                    f"display:flex;align-items:center;gap:12px'>"
                    f"<span style='background:{PAC_TEAL};color:white;border-radius:6px;padding:3px 10px;"
                    f"font-weight:700;font-size:0.8rem;min-width:90px;text-align:center'>{dia}</span>"
                    f"<span style='color:{PAC_NAVY}'>{accion}</span></div>", unsafe_allow_html=True)
            st.caption("Canal principal: llamadas telefónicas + SMS de apoyo. No se utiliza WhatsApp masivo.")

        # ── PAC TAB 5 — REGLAS ──
        with pac5:
            st.markdown("**Reglas por edad → canal asignado**")
            reglas_edad = pd.DataFrame([
                ["18–30","Llamada + SMS — canal digital de respaldo"],
                ["31–45","Llamada + SMS de seguimiento"],
                ["46–60","Llamada prioritaria, reminder por SMS"],
                ["61+",  "Llamada con horario flexible, sin SMS"],
            ], columns=["Rango de edad","Canal asignado"])
            st.dataframe(reglas_edad, use_container_width=True, hide_index=True)

            st.markdown("**Reglas por estado → estrategia**")
            reglas_estado = pd.DataFrame([
                ["Activa","Gestión preventiva, mantener relación comercial"],
                ["Morosa","Gestión de cobranza intensiva según nivel de riesgo"],
            ], columns=["Estado","Estrategia"])
            st.dataframe(reglas_estado, use_container_width=True, hide_index=True)

            st.markdown("**Reglas por segmentación → prioridad**")
            reglas_seg = pd.DataFrame([
                ["Diamante","Prioridad máxima — gestión personalizada"],
                ["Zafiro","Prioridad alta"],
                ["Oro","Prioridad media-alta"],
                ["Plata","Prioridad media"],
                ["Bronce","Prioridad estándar"],
            ], columns=["Segmentación","Prioridad"])
            st.dataframe(reglas_seg, use_container_width=True, hide_index=True)

            st.markdown("---")
            colA, colB = st.columns(2)
            with colA:
                st.markdown("**Distribución por Edad**")
                edad_dist = df_pac["RangoEdad"].value_counts().reindex(orden_edad).fillna(0)
                fig7 = go.Figure(go.Pie(labels=orden_edad, values=edad_dist.values, hole=0.55,
                    marker_colors=[PAC_TEAL,PAC_TEAL2,PAC_AMBER,PAC_CORAL]))
                apply_layout(fig7, height=300)
                st.plotly_chart(fig7, use_container_width=True)
            with colB:
                st.markdown("**Saldo por Segmentación**")
                seg_saldo = df_pac.groupby(c["segmentacion_rep"])[c["saldo_insoluto"]].sum().sort_values(ascending=False)
                fig8 = go.Figure(go.Bar(x=seg_saldo.index, y=seg_saldo.values, marker_color=PAC_CORAL))
                apply_layout(fig8, height=300, yaxis=dict(tickprefix="$", tickformat=",.0f"))
                st.plotly_chart(fig8, use_container_width=True)

        # ── PAC TAB 6 — ZONAS & DIVISIÓN ──
        with pac6:
            st.markdown("**Top 10 zonas por saldo**")
            zonas = df_pac.groupby(c["zona"]).apply(lambda g: pd.Series({
                "Saldo": g[c["saldo_insoluto"]].sum(),
                "Críticas": (g["RiesgoMigracion"]=="Crítico").sum(),
                "Total": len(g),
            })).reset_index().rename(columns={c["zona"]:"Zona"})
            zonas["% Crítico"] = (zonas["Críticas"]/zonas["Total"]*100).round(1)
            zonas = zonas.sort_values("Saldo", ascending=False).head(10)
            st.dataframe(zonas[["Zona","Saldo","% Crítico"]].style.format({"Saldo":"${:,.0f}","% Crítico":"{:.1f}%"}),
                         use_container_width=True, hide_index=True)

            st.markdown("**Top 10 estados de residencia**")
            col_estado_geo_div = pac_col_por_letra(df_pac, PAC_ESTADO_GEO_LETRA)
            if col_estado_geo_div is not None:
                divisiones = df_pac.groupby(col_estado_geo_div)[c["saldo_insoluto"]].sum().sort_values(ascending=False).head(10)
                div_tab = divisiones.reset_index()
                div_tab.columns = ["Estado","Saldo"]
                div_tab["% del total"] = (div_tab["Saldo"]/df_pac[c["saldo_insoluto"]].sum()*100).round(1)
                st.dataframe(div_tab.style.format({"Saldo":"${:,.0f}","% del total":"{:.1f}%"}),
                             use_container_width=True, hide_index=True)
            else:
                st.caption(f"Sube la columna **{PAC_ESTADO_GEO_LETRA}** (direccion_de_residencia_estado) para ver este desglose.")

            st.markdown("---")
            st.markdown("**Cuentas críticas vs total por zona**")
            zonas_chart = zonas.sort_values("Total", ascending=True)
            fig9 = go.Figure()
            fig9.add_trace(go.Bar(name="Total", x=zonas_chart["Total"], y=zonas_chart["Zona"], orientation="h", marker_color="#cbd5e1"))
            fig9.add_trace(go.Bar(name="Críticas", x=zonas_chart["Críticas"], y=zonas_chart["Zona"], orientation="h", marker_color=PAC_RED))
            apply_layout(fig9, height=380, barmode="overlay", legend=dict(orientation="h", y=1.1))

        # ── PAC TAB 7 — SEGUIMIENTO DIARIO DE RECUPERACIÓN ──
        with pac7:
            st.caption(
                "El pago (columna **CF**) es el acumulado de pagos del mes a la fecha del corte cargado, "
                "por lo que la recuperación se calcula directamente sobre ese acumulado — no es necesario subir más de un corte."
            )
            col_pago = c["pago_actual"]
            recuperado_total = df_pac[col_pago].sum()
            n_con_recup = (df_pac[col_pago] > 0).sum()
            n_criticas_sin_recup = ((df_pac["RiesgoMigracion"]=="Crítico") & (df_pac[col_pago]==0)).sum()

            st.markdown("**Recuperación acumulada del mes**")
            k1,k2,k3,k4 = st.columns(4)
            k1.metric("Recuperado acumulado", f"${recuperado_total/1e6:.2f}M")
            k2.metric("Consultoras con recuperación", f"{n_con_recup:,} de {len(df_pac):,}")
            k3.metric("% de cartera que recuperó", f"{n_con_recup/len(df_pac)*100:.1f}%" if len(df_pac) else "0%")
            k4.metric("Críticas SIN recuperación", f"{n_criticas_sin_recup:,}", delta="requieren acción", delta_color="inverse")

            st.markdown("---")
            col_deuda = c["valor_original_deuda"]

            def fig_asignado_vs_recuperado(g, color, orden=None):
                if orden is not None:
                    g = g.reindex(orden).fillna(0)
                pct = (g["Recuperado"]/g["Asignado"]*100).fillna(0)
                fig = go.Figure()
                fig.add_trace(go.Bar(name="Monto asignado", x=g.index.astype(str), y=g["Asignado"], marker_color="#cbd5e1"))
                fig.add_trace(go.Bar(name="Recuperado", x=g.index.astype(str), y=g["Recuperado"], marker_color=color,
                                      text=[f"{p:.1f}%" for p in pct], textposition="outside"))
                apply_layout(fig, height=300, barmode="group", yaxis=dict(tickprefix="$", tickformat=",.0f"),
                             legend=dict(orientation="h", y=1.15))
                return fig

            cdim1, cdim2 = st.columns(2)
            with cdim1:
                st.markdown("**Recuperación por Temporalidad**")
                gt = df_pac.groupby("Temporalidad").agg(Asignado=(col_deuda,"sum"), Recuperado=(col_pago,"sum"))
                st.plotly_chart(fig_asignado_vs_recuperado(gt, PAC_TEAL2, orden_t), use_container_width=True)
            with cdim2:
                st.markdown("**Recuperación por Segmentación (camino de crecimiento)**")
                gs = df_pac.groupby(c["segmentacion_rep"]).agg(Asignado=(col_deuda,"sum"), Recuperado=(col_pago,"sum")).sort_values("Recuperado", ascending=False)
                st.plotly_chart(fig_asignado_vs_recuperado(gs, PAC_AMBER), use_container_width=True)

            cdim3, cdim4 = st.columns(2)
            with cdim3:
                st.markdown("**Recuperación por Estado**")
                col_estado_geo7 = pac_col_por_letra(df_pac, PAC_ESTADO_GEO_LETRA)
                if col_estado_geo7 is not None:
                    ge = df_pac.groupby(col_estado_geo7).agg(Asignado=(col_deuda,"sum"), Recuperado=(col_pago,"sum")).sort_values("Recuperado", ascending=False).head(15)
                    st.plotly_chart(fig_asignado_vs_recuperado(ge, PAC_TEAL), use_container_width=True)
                else:
                    st.caption(f"Sube la columna **{PAC_ESTADO_GEO_LETRA}** (direccion_de_residencia_estado) para ver este desglose.")
            with cdim4:
                st.markdown("**Recuperación por Edad**")
                ga = df_pac.groupby("RangoEdad").agg(Asignado=(col_deuda,"sum"), Recuperado=(col_pago,"sum"))
                st.plotly_chart(fig_asignado_vs_recuperado(ga, PAC_CORAL, orden_edad), use_container_width=True)

            st.markdown("---")
            st.markdown("**Riesgo de migración vs. recuperación — dónde actuar**")
            riesgo_g = df_pac.groupby("RiesgoMigracion").agg(
                Recuperado=(col_pago,"sum"),
                Consultoras=(col_pago,"count"),
                ConRecuperacion=(col_pago, lambda s: (s>0).sum()),
            ).reindex(orden_r).fillna(0)
            riesgo_g["% con recuperación"] = (riesgo_g["ConRecuperacion"]/riesgo_g["Consultoras"]*100).round(1)
            riesgo_tab = riesgo_g.reset_index().rename(columns={"RiesgoMigracion":"Riesgo"})[["Riesgo","Recuperado","Consultoras","% con recuperación"]]
            st.dataframe(
                riesgo_tab.style.format({"Recuperado":"${:,.0f}","% con recuperación":"{:.1f}%"}),
                use_container_width=True, hide_index=True,
            )

            acciones_riesgo = {
                "Crítico":   "Llamada diaria + Llamada de seguimiento + SMS cada 48h + Reminder diario",
                "Alto":      "Llamada 3× por semana + Llamada de seguimiento + SMS 2× por semana",
                "Preventivo":"Llamada 2× por semana + SMS semanal + Reminder semanal",
                "Estable":   "Monitoreo quincenal, sin contacto intensivo",
            }
            sin_recup_por_riesgo = df_pac[df_pac[col_pago]==0]["RiesgoMigracion"].value_counts().reindex(orden_r).fillna(0)
            st.markdown("**Acción recomendada para los grupos que aún no generan recuperación**")
            for riesgo in ["Crítico","Alto","Preventivo"]:
                n_sin = int(sin_recup_por_riesgo.get(riesgo, 0))
                if n_sin > 0:
                    st.markdown(
                        f"<div style='background:{CARD};border-left:4px solid {PAC_RED if riesgo=='Crítico' else (PAC_CORAL if riesgo=='Alto' else PAC_AMBER)};"
                        f"border-radius:10px;padding:12px 16px;margin-bottom:8px'>"
                        f"<b>{riesgo}</b> — {n_sin:,} consultoras sin recuperación todavía. "
                        f"Acción: {acciones_riesgo[riesgo]}.</div>",
                        unsafe_allow_html=True,
                    )

            hist = sorted(st.session_state.pac_history, key=lambda h: h["fecha"])
            if len(hist) >= 2:
                st.markdown("---")
                st.markdown("**Evolución del acumulado entre cortes**")
                filas_hist = []
                for h in hist:
                    filas_hist.append({
                        "Fecha": h["fecha"],
                        "Recuperado acumulado": h["df"][h["cols"]["pago_actual"]].sum(),
                        "Consultoras con recuperación": (h["df"][h["cols"]["pago_actual"]] > 0).sum(),
                    })
                df_hist = pd.DataFrame(filas_hist)
                st.dataframe(df_hist.style.format({"Recuperado acumulado":"${:,.0f}"}), use_container_width=True, hide_index=True)
                fig_hist = go.Figure(go.Bar(x=df_hist["Fecha"].astype(str), y=df_hist["Recuperado acumulado"], marker_color=PAC_TEAL))
                apply_layout(fig_hist, height=280, yaxis=dict(tickprefix="$", tickformat=",.0f"))
                st.plotly_chart(fig_hist, use_container_width=True)
